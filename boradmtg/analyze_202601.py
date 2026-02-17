#!/usr/bin/env python3
"""
Analysis script: Identify major changes in 202601 vs previous months (202506-202512).
Reads サマリー表 from all 8 xlsx files, extracts 合計 column values by row label,
and computes MoM, 3M-avg, 6M-avg deviations.
"""

import openpyxl
import os
import csv
from collections import OrderedDict

BASE_DIR = '/Users/masaaki/Desktop/prm/report/boradmtg/csv/'

# File configuration:
# 202506-202509: labels in cols 2-5, data starts col 6, 合計 at col 32
# 202510-202601: labels in cols 3-6, data starts col 7, 合計 at col 33
FILES = OrderedDict([
    ('202506', {'file': 'デジコン収益分析_202506.xlsx', 'sheet': '2506月サマリー表 (2)',
                'goukei_col': 32, 'label_cols': (2, 6), 'seg_start': 6}),
    ('202507', {'file': 'デジコン収益分析_202507.xlsx', 'sheet': '2507月サマリー表',
                'goukei_col': 32, 'label_cols': (2, 6), 'seg_start': 6}),
    ('202508', {'file': 'デジコン収益分析_202508.xlsx', 'sheet': '2508月サマリー表',
                'goukei_col': 32, 'label_cols': (2, 6), 'seg_start': 6}),
    ('202509', {'file': 'デジコン収益分析_202509.xlsx', 'sheet': '2509月サマリー表',
                'goukei_col': 32, 'label_cols': (2, 6), 'seg_start': 6}),
    ('202510', {'file': 'デジコン収益分析_202510.xlsx', 'sheet': '2510月サマリー表',
                'goukei_col': 33, 'label_cols': (3, 7), 'seg_start': 7}),
    ('202511', {'file': 'デジコン収益分析_202511.xlsx', 'sheet': '2511月サマリー表',
                'goukei_col': 33, 'label_cols': (3, 7), 'seg_start': 7}),
    ('202512', {'file': 'デジコン収益分析_202512.xlsx', 'sheet': '2512月サマリー表 (2)',
                'goukei_col': 33, 'label_cols': (3, 7), 'seg_start': 7}),
    ('202601', {'file': 'デジコン収益分析_202601.xlsx', 'sheet': '2601月サマリー表',
                'goukei_col': 33, 'label_cols': (3, 7), 'seg_start': 7}),
])

# Segment column mappings
# 202510-202601: segments at cols 7-28
SEGMENTS_NEW = {
    7: 'モバイル新規', 8: 'モバイル新規(27)_運用', 9: 'リニューアル(27)_運用',
    10: 'リニューアル(28)候補', 11: '新規_メディア', 12: 'よしもと大集合',
    13: 'モバイル注力', 14: 'はやとも', 15: '注力外',
    16: 'はやともヤースー', 17: '橋本京明', 18: 'めくる/ヤースー',
    19: 'ギャル', 20: 'JUNO', 21: 'Love/sLove',
    22: 'ISP', 23: 'プラットフォーム', 24: 'アプリ',
    25: '海外展開', 26: '自社メディア', 27: '新規_ロト', 28: '新規_ポイント',
}
# 202506-202509: segments at cols 6-27
SEGMENTS_OLD = {
    6: 'モバイル新規', 7: 'モバイル新規(27)_運用', 8: 'リニューアル(27)_運用',
    9: 'リニューアル(28)候補', 10: '新規_メディア', 11: 'よしもと大集合',
    12: 'モバイル注力', 13: 'はやとも', 14: '注力外',
    15: 'はやともヤースー', 16: '橋本京明', 17: 'めくる/ヤースー',
    18: 'ギャル', 19: 'JUNO', 20: 'Love/sLove',
    21: 'ISP', 22: 'プラットフォーム', 23: 'アプリ',
    24: '海外展開', 25: '自社メディア', 26: '新規_ロト', 27: '新規_ポイント',
}


def build_label(ws, row, label_cols):
    """Build a composite label from label columns for a row.
    label_cols = (start, end) where end is exclusive.
    """
    parts = []
    for c in range(label_cols[0], label_cols[1]):
        v = ws.cell(row, c).value
        if v is not None and str(v).strip():
            parts.append(str(v).strip())
    return ' > '.join(parts) if parts else None


def extract_data(period):
    """Extract {row_label: 合計 value} from a file."""
    cfg = FILES[period]
    filepath = os.path.join(BASE_DIR, cfg['file'])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[cfg['sheet']]
    goukei_col = cfg['goukei_col']
    label_cols = cfg['label_cols']

    data = {}
    for r in range(5, ws.max_row + 1):
        label = build_label(ws, r, label_cols)
        if label is None:
            continue
        val = ws.cell(r, goukei_col).value
        if val is not None:
            try:
                val = float(val)
            except (ValueError, TypeError):
                continue
            data[label] = val
    wb.close()
    return data


def extract_segment_data(period, target_labels):
    """Extract segment-level data for specific row labels."""
    cfg = FILES[period]
    filepath = os.path.join(BASE_DIR, cfg['file'])
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[cfg['sheet']]
    label_cols = cfg['label_cols']
    segments = SEGMENTS_OLD if period in ('202506', '202507', '202508', '202509') else SEGMENTS_NEW

    # Map labels to row numbers
    row_label_map = {}
    for r in range(5, ws.max_row + 1):
        label = build_label(ws, r, label_cols)
        if label:
            row_label_map[label] = r

    result = {}
    for tl in target_labels:
        if tl not in row_label_map:
            continue
        row = row_label_map[tl]
        seg_data = {}
        for col, seg_name in segments.items():
            v = ws.cell(row, col).value
            if v is not None:
                try:
                    seg_data[seg_name] = float(v)
                except (ValueError, TypeError):
                    pass
        result[tl] = seg_data
    wb.close()
    return result


def fmt_num(v):
    if v is None:
        return 'N/A'
    return f'{v:,.0f}'


def fmt_pct(v):
    if v is None:
        return 'N/A'
    return f'{v:+.1f}%'


def safe_pct(val, base):
    if base is None or base == 0:
        return None
    return (val - base) / abs(base) * 100


def main():
    print("=" * 120)
    print("デジコン収益分析 202601 変動分析レポート")
    print("=" * 120)

    # ---- Step 1: Extract all data ----
    print("\n[1] Loading all 8 files...")
    all_data = {}
    for period in FILES:
        all_data[period] = extract_data(period)
        print(f"  {period}: {len(all_data[period])} rows extracted")

    ref_labels = list(all_data['202601'].keys())
    print(f"\n  202601 reference: {len(ref_labels)} labeled rows")

    # ---- Step 2: Compute change metrics ----
    print("\n[2] Computing change metrics...")
    results = []
    for label in ref_labels:
        v01 = all_data['202601'].get(label)
        if v01 is None:
            continue
        v12 = all_data['202512'].get(label)

        # 3-month average (202510-202512)
        vals_3m = [all_data[p].get(label) for p in ('202510', '202511', '202512')]
        vals_3m_ok = [v for v in vals_3m if v is not None]
        avg_3m = sum(vals_3m_ok) / len(vals_3m_ok) if vals_3m_ok else None

        # 6-month average (202507-202512)
        vals_6m = [all_data[p].get(label) for p in ('202507', '202508', '202509', '202510', '202511', '202512')]
        vals_6m_ok = [v for v in vals_6m if v is not None]
        avg_6m = sum(vals_6m_ok) / len(vals_6m_ok) if vals_6m_ok else None

        mom_abs = (v01 - v12) if v12 is not None else None
        mom_pct = safe_pct(v01, v12)
        dev_3m = (v01 - avg_3m) if avg_3m is not None else None
        dev_3m_pct = safe_pct(v01, avg_3m)
        dev_6m = (v01 - avg_6m) if avg_6m is not None else None
        dev_6m_pct = safe_pct(v01, avg_6m)

        magnitude = 0
        if mom_abs is not None:
            magnitude += abs(mom_abs)
        if dev_3m is not None:
            magnitude += abs(dev_3m) * 0.5
        if dev_6m is not None:
            magnitude += abs(dev_6m) * 0.3

        results.append({
            'label': label, 'v01': v01, 'v12': v12,
            'mom_abs': mom_abs, 'mom_pct': mom_pct,
            'avg_3m': avg_3m, 'dev_3m': dev_3m, 'dev_3m_pct': dev_3m_pct,
            'avg_6m': avg_6m, 'dev_6m': dev_6m, 'dev_6m_pct': dev_6m_pct,
            'magnitude': magnitude,
        })

    results.sort(key=lambda x: x['magnitude'], reverse=True)

    # ---- Step 3: Print TOP 20 ----
    print("\n" + "=" * 200)
    print("TOP 20 ITEMS BY MAGNITUDE OF CHANGE (202601)")
    print("=" * 200)

    hdr = (f"{'#':>3} | {'Item':<35} | {'202601':>14} | {'202512':>14} | "
           f"{'MoM Chg':>14} | {'MoM%':>8} | {'3M Avg':>14} | {'3M Dev':>14} | "
           f"{'3M%':>8} | {'6M Avg':>14} | {'6M Dev':>14} | {'6M%':>8}")
    print(hdr)
    print("-" * len(hdr))

    for i, r in enumerate(results[:20], 1):
        print(
            f"{i:3d} | {r['label'][:35]:<35} | "
            f"{fmt_num(r['v01']):>14} | {fmt_num(r['v12']):>14} | "
            f"{fmt_num(r['mom_abs']):>14} | {fmt_pct(r['mom_pct']):>8} | "
            f"{fmt_num(r['avg_3m']):>14} | {fmt_num(r['dev_3m']):>14} | "
            f"{fmt_pct(r['dev_3m_pct']):>8} | "
            f"{fmt_num(r['avg_6m']):>14} | {fmt_num(r['dev_6m']):>14} | "
            f"{fmt_pct(r['dev_6m_pct']):>8}"
        )

    # ---- Step 4: Major P&L line items with full trends ----
    print("\n\n" + "=" * 200)
    print("MAJOR P&L LINE ITEMS - FULL MONTHLY TREND (合計)")
    print("=" * 200)

    key_labels_ordered = []
    key_patterns = [
        '売上高',
        '変動費 > 占い師手数料', '決済代行手数料', '支払手数料', '広告宣伝費', '変動費計',
        '限界利益',
        '固定費 > 人件費 > 管理者', '人件費計',
        'その他固定費 > 保守管理費', 'ソフトウェア償却費', '減価償却費', '通信費',
        '外注加工費', '業務委託料', '賃借料', '原稿料', '採用費',
        'その他固定費計',
        '仕掛品 > 取崩額', '計上額', '仕掛品計',
        '固定費計',
        'プロモーション配賦', 'その他配賦', '営業利益',
    ]
    for kp in key_patterns:
        for label in ref_labels:
            if label == kp or label.endswith(kp) or kp == label.split(' > ')[-1]:
                if label not in key_labels_ordered:
                    key_labels_ordered.append(label)
                    break
        else:
            # try partial match
            for label in ref_labels:
                if kp in label and label not in key_labels_ordered:
                    key_labels_ordered.append(label)
                    break

    periods = list(FILES.keys())
    ph = f"  {'Item':<40} | " + " | ".join(f"{p:>14}" for p in periods) + f" | {'MoM Chg':>14} | {'MoM%':>8}"
    print(ph)
    print("  " + "-" * (len(ph) - 2))

    for label in key_labels_ordered:
        vals = []
        for p in periods:
            v = all_data[p].get(label)
            vals.append(fmt_num(v) if v is not None else '-')

        v01 = all_data['202601'].get(label)
        v12 = all_data['202512'].get(label)
        mom = (v01 - v12) if (v01 is not None and v12 is not None) else None
        pct = safe_pct(v01, v12) if (v01 is not None and v12 is not None and v12 != 0) else None

        print(f"  {label[:40]:<40} | " + " | ".join(f"{v:>14}" for v in vals) +
              f" | {fmt_num(mom):>14} | {fmt_pct(pct):>8}")

    # ---- Step 5: Segment-level analysis ----
    print("\n\n" + "=" * 200)
    print("SEGMENT-LEVEL MoM ANALYSIS (202601 vs 202512)")
    print("=" * 200)

    seg_target = ['売上高', '変動費計', '限界利益', '人件費計', 'その他固定費計', '仕掛品計', '固定費計', '営業利益']
    seg_data_01 = extract_segment_data('202601', seg_target)
    seg_data_12 = extract_segment_data('202512', seg_target)
    seg_names = list(SEGMENTS_NEW.values())

    for tl in seg_target:
        if tl not in seg_data_01:
            continue
        print(f"\n--- {tl} ---")
        s01 = seg_data_01.get(tl, {})
        s12 = seg_data_12.get(tl, {})

        changes = []
        for sn in seg_names:
            a = s01.get(sn, 0) or 0
            b = s12.get(sn, 0) or 0
            chg = a - b
            pct = safe_pct(a, b) if b != 0 else None
            changes.append((sn, a, b, chg, pct))
        changes.sort(key=lambda x: abs(x[3]), reverse=True)

        sh = f"  {'Segment':<28} | {'202601':>14} | {'202512':>14} | {'Change':>14} | {'%':>8}"
        print(sh)
        print("  " + "-" * (len(sh) - 2))
        for sn, a, b, chg, pct in changes:
            if chg == 0 and a == 0:
                continue
            print(f"  {sn:<28} | {fmt_num(a):>14} | {fmt_num(b):>14} | {fmt_num(chg):>14} | {fmt_pct(pct):>8}")

    # ---- Step 6: Save CSV ----
    output_path = '/Users/masaaki/Desktop/prm/report/boradmtg/analysis_202601.csv'
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)

        w.writerow(['=== 変動ランキング (全項目) ==='])
        w.writerow(['Rank', 'Item', '202601', '202512', 'MoM Change', 'MoM %',
                     '3M Avg (202510-12)', '3M Deviation', '3M %',
                     '6M Avg (202507-12)', '6M Deviation', '6M %', 'Magnitude'])
        for i, r in enumerate(results, 1):
            w.writerow([
                i, r['label'], r['v01'], r['v12'],
                r['mom_abs'],
                f"{r['mom_pct']:.1f}%" if r['mom_pct'] is not None else 'N/A',
                r['avg_3m'], r['dev_3m'],
                f"{r['dev_3m_pct']:.1f}%" if r['dev_3m_pct'] is not None else 'N/A',
                r['avg_6m'], r['dev_6m'],
                f"{r['dev_6m_pct']:.1f}%" if r['dev_6m_pct'] is not None else 'N/A',
                r['magnitude']
            ])

        w.writerow([])
        w.writerow(['=== 主要P&L月次推移 ==='])
        w.writerow(['Item'] + list(periods))
        for label in key_labels_ordered:
            row = [label]
            for p in periods:
                row.append(all_data[p].get(label, ''))
            w.writerow(row)

        w.writerow([])
        w.writerow(['=== セグメント別MoM分析 ==='])
        for tl in seg_target:
            if tl not in seg_data_01:
                continue
            w.writerow([f'--- {tl} ---'])
            w.writerow(['Segment', '202601', '202512', 'Change', '%'])
            s01 = seg_data_01.get(tl, {})
            s12 = seg_data_12.get(tl, {})
            changes = []
            for sn in seg_names:
                a = s01.get(sn, 0) or 0
                b = s12.get(sn, 0) or 0
                chg = a - b
                pct = safe_pct(a, b) if b != 0 else None
                changes.append((sn, a, b, chg, pct))
            changes.sort(key=lambda x: abs(x[3]), reverse=True)
            for sn, a, b, chg, pct in changes:
                w.writerow([sn, a, b, chg, f"{pct:.1f}%" if pct is not None else 'N/A'])

    print(f"\n\nFull analysis saved to: {output_path}")
    print("=" * 120)


if __name__ == '__main__':
    main()
