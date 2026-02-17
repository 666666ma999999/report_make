#!/usr/bin/env python3
"""
Board-level monthly report generator for January 2026.
Generates 月次報告_202601.xlsx with 4 sheets:
  1. エグゼクティブサマリー
  2. セグメント分析
  3. トレンド推移
  4. 全行詳細データ
"""

import pandas as pd
import numpy as np
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ===== Configuration =====
BASE_DIR = "/Users/masaaki/Desktop/prm/report/boradmtg"
CSV_DIR = os.path.join(BASE_DIR, "csv")
OUTPUT_FILE = os.path.join(BASE_DIR, "月次報告_202601.xlsx")
ANALYSIS_CSV = os.path.join(BASE_DIR, "analysis_202601.csv")
SEGMENT_CSV = os.path.join(BASE_DIR, "segment_analysis_202601.csv")

MONTHS = ["202506", "202507", "202508", "202509", "202510", "202511", "202512", "202601"]
MONTH_LABELS = ["25/06", "25/07", "25/08", "25/09", "25/10", "25/11", "25/12", "26/01"]

# Colors
DARK_BLUE = "1F4E79"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN_BG = "C6EFCE"
RED_BG = "FFC7CE"
DARK_GRAY = "404040"
SECTION_GRAY = "595959"

# Fonts
TITLE_FONT = Font(name="Yu Gothic UI", size=14, bold=True)
SECTION_FONT = Font(name="Yu Gothic UI", size=11, bold=True)
HEADER_FONT = Font(name="Yu Gothic UI", size=10, bold=True, color=WHITE)
DATA_FONT = Font(name="Yu Gothic UI", size=10)
DATA_FONT_RED = Font(name="Yu Gothic UI", size=10, color="FF0000")
DATA_FONT_BOLD_RED = Font(name="Yu Gothic UI", size=10, bold=True, color="FF0000")
DATA_FONT_GREEN = Font(name="Yu Gothic UI", size=10, color="006100")
SMALL_FONT = Font(name="Yu Gothic UI", size=9)

# Fills
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
GREEN_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
RED_FILL = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
SECTION_FILL = PatternFill(start_color=SECTION_GRAY, end_color=SECTION_GRAY, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Alignment
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

# Number formats
NUM_FMT_CURRENCY = '#,##0'
NUM_FMT_PCT = '0.0%'
NUM_FMT_PCT_DISPLAY = '0.0"%"'


# ===== Data Loading =====

def load_xlsx_data(filepath):
    """Load a single xlsx file and return raw DataFrame."""
    df = pd.read_excel(filepath, header=None, engine="openpyxl")
    return df


def find_row_by_label(df, label, label_cols=None):
    """Find a row by matching label text in specified columns."""
    if label_cols is None:
        label_cols = [1, 2, 3, 4]  # columns B-E (0-indexed)
    for idx, row in df.iterrows():
        for col in label_cols:
            if col < len(row):
                cell_val = str(row.iloc[col]).strip() if pd.notna(row.iloc[col]) else ""
                if cell_val == label:
                    return idx
    return None


def find_header_row(df):
    """Find the row with segment names (header row)."""
    for idx, row in df.iterrows():
        row_str = " ".join([str(v) for v in row if pd.notna(v)])
        if "モバイル新規" in row_str and "はやとも" in row_str:
            return idx
    return None


def extract_totals_column(df, row_idx):
    """Extract the value from the last (合計) column."""
    if row_idx is None:
        return 0
    row = df.iloc[row_idx]
    # Find last non-NaN numeric value
    for col_idx in range(len(row) - 1, -1, -1):
        val = row.iloc[col_idx]
        if pd.notna(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                continue
    return 0


def extract_all_segment_values(df, row_idx, header_row_idx):
    """Extract all segment values from a row."""
    if row_idx is None:
        return {}
    segments = {}
    header = df.iloc[header_row_idx]
    row = df.iloc[row_idx]
    for col_idx in range(5, len(row)):
        seg_name = str(header.iloc[col_idx]).strip() if col_idx < len(header) and pd.notna(header.iloc[col_idx]) else None
        if seg_name and seg_name != "nan" and seg_name != "合計":
            try:
                val = float(row.iloc[col_idx]) if pd.notna(row.iloc[col_idx]) else 0
            except (ValueError, TypeError):
                val = 0
            segments[seg_name] = val
    return segments


# Key P&L row labels to search for
PL_ITEMS_MAIN = [
    "売上高", "変動費計", "限界利益", "人件費計", "固定費計", "営業利益"
]

PL_ITEMS_DETAIL = [
    "売上高",
    "占い師手数料", "決済代行手数料", "支払手数料", "広告宣伝費", "変動費計",
    "限界利益",
    "人件費計",
    "保守管理費", "ソフトウェア償却費", "減価償却費", "通信費",
    "外注加工費", "業務委託料", "賃借料", "原稿料", "採用費",
    "その他(固定費)", "仕掛品取崩額", "仕掛品計上額", "仕掛品計", "固定費計",
    "プロモーション配賦", "その他配賦", "営業利益"
]

# Alternate label mappings for search
LABEL_ALIASES = {
    "占い師手数料": ["占い師手数料"],
    "支払手数料": ["支払手数料", "（LINE@）"],
    "その他(固定費)": ["その他固定費計", "その他(固定費)"],
    "仕掛品取崩額": ["取崩額"],
    "仕掛品計上額": ["計上額"],
    "仕掛品計": ["仕掛品計"],
}


def parse_analysis_csv():
    """Parse the analysis CSV file for rankings and trend data."""
    data = {
        "rankings": [],
        "trends": {},
        "segment_mom": {},
        "flagged_segments": [],
    }

    current_section = None
    current_subsection = None

    with open(ANALYSIS_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row_data in reader:
            if not row_data or all(c.strip() == "" for c in row_data):
                continue

            line = row_data[0] if row_data else ""

            if "=== 変動ランキング" in line:
                current_section = "ranking"
                continue
            elif "=== 主要P&L月次推移" in line:
                current_section = "trend"
                continue
            elif "=== セグメント別MoM分析" in line:
                current_section = "segment_mom"
                continue
            elif "=== Flagged Segments" in line:
                current_section = "flags"
                continue
            elif line.startswith("---"):
                current_subsection = line.strip("- ")
                continue

            if current_section == "ranking":
                if row_data[0] == "Rank":
                    continue  # header
                try:
                    rank_entry = {
                        "rank": int(row_data[0]),
                        "item": row_data[1],
                        "val_202601": safe_float(row_data[2]),
                        "val_202512": safe_float(row_data[3]),
                        "mom_change": safe_float(row_data[4]),
                        "mom_pct": row_data[5],
                        "avg_3m": safe_float(row_data[6]),
                        "dev_3m": safe_float(row_data[7]),
                        "pct_3m": row_data[8],
                        "avg_6m": safe_float(row_data[9]),
                        "dev_6m": safe_float(row_data[10]),
                        "pct_6m": row_data[11],
                        "magnitude": safe_float(row_data[12]) if len(row_data) > 12 else 0,
                    }
                    data["rankings"].append(rank_entry)
                except (IndexError, ValueError):
                    pass

            elif current_section == "trend":
                if row_data[0] == "Item":
                    continue
                try:
                    item_name = row_data[0]
                    values = [safe_float(v) for v in row_data[1:9]]
                    data["trends"][item_name] = values
                except (IndexError, ValueError):
                    pass

            elif current_section == "flags":
                if row_data[0] == "Segment":
                    continue
                try:
                    data["flagged_segments"].append({
                        "segment": row_data[0],
                        "flag": row_data[1] if len(row_data) > 1 else "",
                    })
                except (IndexError, ValueError):
                    pass

    return data


def parse_segment_csv():
    """Parse the segment analysis CSV.
    Columns (20 total):
    0: Segment, 1: P&L Item, 2: 202601 Value, 3: MoM Change (abs), 4: MoM Change (%),
    5: 3M Avg, 6: 3M Deviation (abs), 7: 3M Deviation (%),
    8: 6M Avg, 9: 6M Deviation (abs), 10: 6M Deviation (%),
    11-18: Trend_202506..Trend_202601, 19: Flags
    """
    segments = {}

    with open(SEGMENT_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)  # skip header

        for row_data in reader:
            if not row_data or len(row_data) < 11:
                continue
            # Skip section separators
            if row_data[0].startswith("===") or row_data[0].startswith("---"):
                continue

            seg_name = row_data[0].strip()
            pl_item = row_data[1].strip()

            if not seg_name or not pl_item:
                continue

            if seg_name not in segments:
                segments[seg_name] = {}

            segments[seg_name][pl_item] = {
                "val_202601": safe_float(row_data[2]) if len(row_data) > 2 else 0,
                "mom_change": safe_float(row_data[3]) if len(row_data) > 3 else 0,
                "mom_pct": safe_float(row_data[4]) if len(row_data) > 4 else 0,
                "avg_3m": safe_float(row_data[5]) if len(row_data) > 5 else 0,
                "dev_3m": safe_float(row_data[6]) if len(row_data) > 6 else 0,
                "pct_3m": safe_float(row_data[7]) if len(row_data) > 7 else 0,
                "avg_6m": safe_float(row_data[8]) if len(row_data) > 8 else 0,
                "dev_6m": safe_float(row_data[9]) if len(row_data) > 9 else 0,
                "pct_6m": safe_float(row_data[10]) if len(row_data) > 10 else 0,
                "trends": [safe_float(row_data[i]) for i in range(11, min(19, len(row_data)))] if len(row_data) > 11 else [],
                "flags": row_data[19] if len(row_data) > 19 else "",
            }

    return segments


def safe_float(val):
    """Safely convert to float."""
    if val is None or val == "" or val == "N/A" or val == "nan":
        return 0.0
    try:
        s = str(val).replace("%", "").replace(",", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return 0.0


# ===== Helper Functions =====

def format_number(val, is_pct=False):
    """Format number for display."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "-"
    if is_pct:
        return f"{val:.1f}%"
    if abs(val) >= 1:
        return f"{val:,.0f}"
    return f"{val:.2f}"


def apply_header_style(ws, row, start_col, end_col):
    """Apply header styling to a row."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def apply_data_style(ws, row, start_col, end_col, is_alt=False):
    """Apply data styling to a row."""
    fill = ALT_ROW_FILL if is_alt else WHITE_FILL
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill = fill


def apply_conditional_fill(cell, val, threshold=0):
    """Apply green/red fill based on value."""
    try:
        v = float(val) if val is not None else 0
    except (ValueError, TypeError):
        return
    if v > threshold:
        cell.fill = GREEN_FILL
        cell.font = DATA_FONT_GREEN
    elif v < -threshold:
        cell.fill = RED_FILL
        cell.font = DATA_FONT_RED


def write_section_header(ws, row, col, text, merge_end_col=None):
    """Write a section header."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Yu Gothic UI", size=11, bold=True, color=WHITE)
    cell.fill = SECTION_FILL
    cell.alignment = LEFT_ALIGN
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER


def write_value_cell(ws, row, col, value, is_currency=True, is_pct=False, is_alt=False):
    """Write a formatted value cell."""
    cell = ws.cell(row=row, column=col)
    fill = ALT_ROW_FILL if is_alt else WHITE_FILL

    if value is None or (isinstance(value, float) and (np.isnan(value) or np.isinf(value))):
        cell.value = "-"
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = fill
        cell.border = THIN_BORDER
        return cell

    cell.value = value
    cell.alignment = RIGHT_ALIGN
    cell.border = THIN_BORDER
    cell.fill = fill

    if is_pct:
        cell.number_format = '0.0%'
        cell.value = value / 100.0 if abs(value) < 10 and not is_currency else value / 100.0
    elif is_currency:
        cell.number_format = NUM_FMT_CURRENCY

    # Negative number formatting
    if isinstance(value, (int, float)) and not is_pct:
        if value < 0:
            cell.font = DATA_FONT_RED
        else:
            cell.font = DATA_FONT

    return cell


def generate_comment(item, val_cur, val_prev, mom_pct, pct_3m, pct_6m):
    """Generate AI-style draft comment in Japanese for a change."""
    comments = []
    item_str = str(item)

    mom_pct_val = safe_float(str(mom_pct).replace("%", ""))
    pct_3m_val = safe_float(str(pct_3m).replace("%", ""))
    pct_6m_val = safe_float(str(pct_6m).replace("%", ""))

    # Direction
    if mom_pct_val > 0:
        direction = "増加"
        sign = "+"
    elif mom_pct_val < 0:
        direction = "減少"
        sign = ""
    else:
        return "前月と同水準"

    # Item-specific comments
    if "広告宣伝費" in item_str:
        if mom_pct_val > 10:
            comments.append(f"新規獲得施策の集中投下により前月比{sign}{mom_pct_val:.1f}%")
            if abs(pct_3m_val) > 20:
                comments.append(f"3M平均比{pct_3m_val:+.1f}%と大幅増")
            comments.append("費用対効果の検証が必要")
        elif mom_pct_val < -10:
            comments.append(f"広告投下量を抑制、前月比{mom_pct_val:.1f}%")
        else:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%で微{direction}")
    elif "売上高" in item_str:
        if mom_pct_val > 0:
            comments.append(f"はやとも・Love/sLove好調により前月比{sign}{mom_pct_val:.1f}%")
        else:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
        if abs(pct_6m_val) > 20:
            comments.append(f"6M平均比{pct_6m_val:+.1f}%、中期トレンドに注意")
    elif "営業利益" in item_str:
        comments.append(f"売上は+5.5%成長も広告費+22.6%増が利益を圧迫")
        if pct_6m_val < -20:
            comments.append(f"6M平均比{pct_6m_val:.1f}%と構造的な利益率低下傾向")
        elif pct_6m_val > 20:
            comments.append(f"6M平均比+{pct_6m_val:.1f}%と改善傾向")
    elif "限界利益" in item_str:
        comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
        if mom_pct_val < 0:
            comments.append("売上+5.5%成長も変動費+16.2%増により限界利益率が悪化(54.3%→49.6%)")
        elif mom_pct_val > 5:
            comments.append("変動費抑制により限界利益率が改善")
    elif "変動費" in item_str:
        comments.append(f"前月比{sign}{mom_pct_val:.1f}%の{direction}")
        if abs(pct_3m_val) > 20:
            comments.append(f"3M平均比{pct_3m_val:+.1f}%")
        if mom_pct_val > 10:
            comments.append("広告宣伝費の急増が主因。売上成長率を上回るコスト増に注意")
    elif "人件費" in item_str:
        comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
        if abs(mom_pct_val) > 5:
            if mom_pct_val > 0:
                comments.append("人員増強・賞与等の影響")
            else:
                comments.append("退職・異動等による変動")
    elif "固定費" in item_str:
        comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
    elif "仕掛品" in item_str:
        if abs(mom_pct_val) > 50:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%と大幅変動")
            comments.append("開発案件の進捗状況を確認")
        else:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
    elif "齋藤勇" in item_str or "高橋明日香" in item_str or "村上朋子" in item_str:
        if val_cur == 0 and val_prev > 0:
            comments.append("当月計上なし（退職・異動等）")
        elif val_cur > 0 and val_prev == 0:
            comments.append("当月新規計上（配属・採用等）")
        else:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
    elif "佐藤花菜子" in item_str:
        comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
        if mom_pct_val > 50:
            comments.append("勤務形態変更・昇給等の影響")
    elif "成田元絵" in item_str:
        comments.append(f"前月比{mom_pct_val:.1f}%")
        if mom_pct_val < -15:
            comments.append("勤務時間減少等")
    elif "決済代行" in item_str:
        comments.append(f"売上連動で前月比{sign}{mom_pct_val:.1f}%")
    elif "減価償却" in item_str:
        comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
        if mom_pct_val > 20:
            comments.append("新規資産取得による増加")
    elif "原稿料" in item_str:
        comments.append(f"前月比{mom_pct_val:.1f}%")
        if mom_pct_val < -20:
            comments.append("コンテンツ制作の抑制")
    else:
        # Generic
        if abs(mom_pct_val) > 50:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%と大幅{direction}")
        elif abs(mom_pct_val) > 10:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%")
        else:
            comments.append(f"前月比{sign}{mom_pct_val:.1f}%で安定推移")

    return "。".join(comments)


# ===== Sheet Builders =====

def build_sheet1_executive_summary(wb, analysis_data, segment_data):
    """Build Sheet 1: エグゼクティブサマリー"""
    ws = wb.active
    ws.title = "エグゼクティブサマリー"

    # Page setup
    ws.sheet_properties.pageSetUpPr = None

    # --- Title ---
    ws.merge_cells("A1:J1")
    cell = ws.cell(row=1, column=1, value="デジコン事業 月次報告書 2026年1月")
    cell.font = TITLE_FONT
    cell.alignment = LEFT_ALIGN

    ws.cell(row=2, column=1, value="報告日: 2026年2月").font = Font(name="Yu Gothic UI", size=10)

    # --- 主要トピックス ---
    write_section_header(ws, 5, 1, "■ 主要トピックス", 10)
    ws.cell(row=6, column=1, value="(後日記入)").font = Font(name="Yu Gothic UI", size=10, italic=True, color="808080")

    # --- 主要KPI ---
    write_section_header(ws, 9, 1, "■ 主要KPI", 10)

    # KPI Headers
    kpi_headers = ["項目", "当月(202601)", "前月(202512)", "前月比", "前月比%", "3M平均", "3M乖離%", "6M平均", "6M乖離%"]
    for i, h in enumerate(kpi_headers, 1):
        apply_header_style(ws, 10, 1, len(kpi_headers))
        ws.cell(row=10, column=i, value=h)

    # KPI Data
    kpi_items = ["売上高", "変動費計", "限界利益", "限界利益率", "人件費計", "固定費計", "営業利益", "営業利益率"]
    trends = analysis_data["trends"]

    for idx, item in enumerate(kpi_items):
        row = 11 + idx
        is_alt = idx % 2 == 1

        if item == "限界利益率":
            # Calculate from 限界利益 / 売上高
            rev_cur = trends.get("売上高", [0]*8)[7]
            rev_prev = trends.get("売上高", [0]*8)[6]
            mp_cur = trends.get("限界利益", [0]*8)[7]
            mp_prev = trends.get("限界利益", [0]*8)[6]

            rate_cur = (mp_cur / rev_cur * 100) if rev_cur else 0
            rate_prev = (mp_prev / rev_prev * 100) if rev_prev else 0
            rate_diff = rate_cur - rate_prev

            # 3M avg
            mp_3m = np.mean(trends.get("限界利益", [0]*8)[4:7])
            rev_3m = np.mean(trends.get("売上高", [0]*8)[4:7])
            rate_3m = (mp_3m / rev_3m * 100) if rev_3m else 0

            # 6M avg
            mp_6m = np.mean(trends.get("限界利益", [0]*8)[1:7])
            rev_6m = np.mean(trends.get("売上高", [0]*8)[1:7])
            rate_6m = (mp_6m / rev_6m * 100) if rev_6m else 0

            ws.cell(row=row, column=1, value=item).font = DATA_FONT
            c2 = ws.cell(row=row, column=2, value=rate_cur / 100)
            c2.number_format = '0.0%'
            c3 = ws.cell(row=row, column=3, value=rate_prev / 100)
            c3.number_format = '0.0%'
            c4 = ws.cell(row=row, column=4, value=rate_diff / 100)
            c4.number_format = '+0.0%;-0.0%'
            ws.cell(row=row, column=5, value="-").alignment = CENTER_ALIGN
            c6 = ws.cell(row=row, column=6, value=rate_3m / 100)
            c6.number_format = '0.0%'
            c7 = ws.cell(row=row, column=7, value=(rate_cur - rate_3m) / 100)
            c7.number_format = '+0.0%;-0.0%'
            c8 = ws.cell(row=row, column=8, value=rate_6m / 100)
            c8.number_format = '0.0%'
            c9 = ws.cell(row=row, column=9, value=(rate_cur - rate_6m) / 100)
            c9.number_format = '+0.0%;-0.0%'

            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
                if col >= 2:
                    cell.alignment = RIGHT_ALIGN
                    cell.font = DATA_FONT

            # Conditional fill for change columns
            apply_conditional_fill(ws.cell(row=row, column=4), rate_diff)
            apply_conditional_fill(ws.cell(row=row, column=7), rate_cur - rate_3m)
            apply_conditional_fill(ws.cell(row=row, column=9), rate_cur - rate_6m)
            continue

        elif item == "営業利益率":
            rev_cur = trends.get("売上高", [0]*8)[7]
            rev_prev = trends.get("売上高", [0]*8)[6]
            op_cur = trends.get("営業利益", [0]*8)[7]
            op_prev = trends.get("営業利益", [0]*8)[6]

            rate_cur = (op_cur / rev_cur * 100) if rev_cur else 0
            rate_prev = (op_prev / rev_prev * 100) if rev_prev else 0
            rate_diff = rate_cur - rate_prev

            op_3m = np.mean(trends.get("営業利益", [0]*8)[4:7])
            rev_3m = np.mean(trends.get("売上高", [0]*8)[4:7])
            rate_3m = (op_3m / rev_3m * 100) if rev_3m else 0

            op_6m = np.mean(trends.get("営業利益", [0]*8)[1:7])
            rev_6m = np.mean(trends.get("売上高", [0]*8)[1:7])
            rate_6m = (op_6m / rev_6m * 100) if rev_6m else 0

            ws.cell(row=row, column=1, value=item).font = DATA_FONT
            c2 = ws.cell(row=row, column=2, value=rate_cur / 100)
            c2.number_format = '0.0%'
            c3 = ws.cell(row=row, column=3, value=rate_prev / 100)
            c3.number_format = '0.0%'
            c4 = ws.cell(row=row, column=4, value=rate_diff / 100)
            c4.number_format = '+0.0%;-0.0%'
            ws.cell(row=row, column=5, value="-").alignment = CENTER_ALIGN
            c6 = ws.cell(row=row, column=6, value=rate_3m / 100)
            c6.number_format = '0.0%'
            c7 = ws.cell(row=row, column=7, value=(rate_cur - rate_3m) / 100)
            c7.number_format = '+0.0%;-0.0%'
            c8 = ws.cell(row=row, column=8, value=rate_6m / 100)
            c8.number_format = '0.0%'
            c9 = ws.cell(row=row, column=9, value=(rate_cur - rate_6m) / 100)
            c9.number_format = '+0.0%;-0.0%'

            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
                if col >= 2:
                    cell.alignment = RIGHT_ALIGN
                    cell.font = DATA_FONT

            apply_conditional_fill(ws.cell(row=row, column=4), rate_diff)
            apply_conditional_fill(ws.cell(row=row, column=7), rate_cur - rate_3m)
            apply_conditional_fill(ws.cell(row=row, column=9), rate_cur - rate_6m)
            continue

        # Regular items
        trend_vals = trends.get(item, [0]*8)
        val_cur = trend_vals[7] if len(trend_vals) > 7 else 0
        val_prev = trend_vals[6] if len(trend_vals) > 6 else 0
        mom_change = val_cur - val_prev
        mom_pct = (mom_change / abs(val_prev) * 100) if val_prev != 0 else 0

        # 3M average (202510-202512)
        avg_3m = np.mean(trend_vals[4:7]) if len(trend_vals) >= 7 else 0
        dev_3m_pct = ((val_cur - avg_3m) / abs(avg_3m) * 100) if avg_3m != 0 else 0

        # 6M average (202507-202512)
        avg_6m = np.mean(trend_vals[1:7]) if len(trend_vals) >= 7 else 0
        dev_6m_pct = ((val_cur - avg_6m) / abs(avg_6m) * 100) if avg_6m != 0 else 0

        ws.cell(row=row, column=1, value=item).font = DATA_FONT

        c2 = ws.cell(row=row, column=2, value=val_cur)
        c2.number_format = NUM_FMT_CURRENCY
        c3 = ws.cell(row=row, column=3, value=val_prev)
        c3.number_format = NUM_FMT_CURRENCY
        c4 = ws.cell(row=row, column=4, value=mom_change)
        c4.number_format = NUM_FMT_CURRENCY
        c5 = ws.cell(row=row, column=5, value=mom_pct / 100)
        c5.number_format = '+0.0%;-0.0%'
        c6 = ws.cell(row=row, column=6, value=avg_3m)
        c6.number_format = NUM_FMT_CURRENCY
        c7 = ws.cell(row=row, column=7, value=dev_3m_pct / 100)
        c7.number_format = '+0.0%;-0.0%'
        c8 = ws.cell(row=row, column=8, value=avg_6m)
        c8.number_format = NUM_FMT_CURRENCY
        c9 = ws.cell(row=row, column=9, value=dev_6m_pct / 100)
        c9.number_format = '+0.0%;-0.0%'

        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            if col >= 2:
                cell.alignment = RIGHT_ALIGN

        # Color-code change columns
        apply_conditional_fill(ws.cell(row=row, column=4), mom_change if "費" in item else -mom_change)
        apply_conditional_fill(ws.cell(row=row, column=5), mom_pct if item in ["売上高", "限界利益", "営業利益"] else -mom_pct)

    # --- 変動アラート TOP10 ---
    write_section_header(ws, 20, 1, "■ 変動アラート TOP10", 10)

    alert_headers = ["No.", "区分", "項目", "当月", "前月", "前月比", "前月比%", "3M平均乖離%", "6M平均乖離%", "コメント"]
    for i, h in enumerate(alert_headers, 1):
        apply_header_style(ws, 21, 1, len(alert_headers))
        ws.cell(row=21, column=i, value=h)

    # Get top 10 from rankings
    rankings = analysis_data["rankings"][:10]
    for idx, entry in enumerate(rankings):
        row = 22 + idx
        is_alt = idx % 2 == 1

        # Determine category
        item_name = entry["item"]
        if item_name in ["売上高", "限界利益", "営業利益"]:
            category = "損益"
        elif "費" in item_name or "手数料" in item_name:
            category = "コスト"
        elif "仕掛品" in item_name or "計上額" in item_name or "取崩額" in item_name:
            category = "仕掛品"
        else:
            category = "人件費"

        comment = generate_comment(
            item_name, entry["val_202601"], entry["val_202512"],
            entry["mom_pct"], entry["pct_3m"], entry["pct_6m"]
        )

        ws.cell(row=row, column=1, value=idx + 1).font = DATA_FONT
        ws.cell(row=row, column=2, value=category).font = DATA_FONT
        ws.cell(row=row, column=3, value=item_name).font = DATA_FONT
        c4 = ws.cell(row=row, column=4, value=entry["val_202601"])
        c4.number_format = NUM_FMT_CURRENCY
        c5 = ws.cell(row=row, column=5, value=entry["val_202512"])
        c5.number_format = NUM_FMT_CURRENCY
        c6 = ws.cell(row=row, column=6, value=entry["mom_change"])
        c6.number_format = NUM_FMT_CURRENCY

        mom_pct_val = safe_float(str(entry["mom_pct"]).replace("%", ""))
        c7 = ws.cell(row=row, column=7, value=mom_pct_val / 100)
        c7.number_format = '+0.0%;-0.0%'

        pct_3m_val = safe_float(str(entry["pct_3m"]).replace("%", ""))
        c8 = ws.cell(row=row, column=8, value=pct_3m_val / 100)
        c8.number_format = '+0.0%;-0.0%'

        pct_6m_val = safe_float(str(entry["pct_6m"]).replace("%", ""))
        c9 = ws.cell(row=row, column=9, value=pct_6m_val / 100)
        c9.number_format = '+0.0%;-0.0%'

        ws.cell(row=row, column=10, value=comment).font = SMALL_FONT

        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            if col >= 4 and col <= 9:
                cell.alignment = RIGHT_ALIGN

        ws.cell(row=row, column=10).alignment = LEFT_ALIGN

        # Conditional formatting on MoM%
        if mom_pct_val > 10:
            if category == "コスト":
                ws.cell(row=row, column=7).fill = RED_FILL
            else:
                ws.cell(row=row, column=7).fill = GREEN_FILL
        elif mom_pct_val < -10:
            if category in ["損益"]:
                ws.cell(row=row, column=7).fill = RED_FILL
            else:
                ws.cell(row=row, column=7).fill = GREEN_FILL

    # --- 要注意セグメント ---
    write_section_header(ws, 33, 1, "■ 要注意セグメント", 10)

    flag_headers = ["セグメント", "フラグ", "当月営業利益", "前月営業利益", "変動", "詳細"]
    for i, h in enumerate(flag_headers, 1):
        apply_header_style(ws, 34, 1, len(flag_headers))
        ws.cell(row=34, column=i, value=h)

    # Key flagged segments
    flagged_items = [
        {
            "segment": "はやともヤースー",
            "flag": "黒字→赤字転落",
            "detail": "変動費+40.4%が売上+16.1%を大幅に上回り、営業利益が黒字(765,564)から赤字(-477,840)に転落。変動費率の急上昇が主因。"
        },
        {
            "segment": "新規_メディア(大百科)",
            "flag": "赤字拡大・変動費急増",
            "detail": "変動費+123.2%と急増し、営業利益が-77.5万円に悪化。3M平均比-1337%と大幅乖離。事業モデルの見直しが必要。"
        },
        {
            "segment": "モバイル新規",
            "flag": "売上-29.2%",
            "detail": "売上高が前月比-29.2%と大幅減。6M平均比-44.9%。新規獲得チャネルの見直しが急務。"
        },
        {
            "segment": "海外展開",
            "flag": "構造的下落",
            "detail": "売上-15.1%、3M平均比-56.1%、6M平均比-58.9%。継続的な下落トレンド。事業継続判断が必要。"
        },
        {
            "segment": "リニューアル(28)候補",
            "flag": "固定費急増",
            "detail": "固定費+455.1%(43,481→241,349)。営業利益は-48.9%と半減。投資回収計画の再検証が必要。"
        },
        {
            "segment": "自社メディア",
            "flag": "利益率悪化",
            "detail": "売上-2.0%に対しコスト+6.9%。営業利益は-27.5%と3M/6M平均比-25.4%。コスト構造の改善が課題。"
        },
    ]

    for idx, item in enumerate(flagged_items):
        row = 35 + idx
        is_alt = idx % 2 == 1

        seg_name = item["segment"]
        # Map segment name for lookup
        seg_lookup = seg_name
        if seg_name == "新規_メディア(大百科)":
            seg_lookup = "新規_メディア(大百科)"

        seg_d = segment_data.get(seg_lookup, {})
        op_data = seg_d.get("営業利益", {})
        op_cur = op_data.get("val_202601", 0) if op_data else 0
        op_prev = op_cur - op_data.get("mom_change", 0) if op_data else 0
        op_change = op_data.get("mom_change", 0) if op_data else 0

        ws.cell(row=row, column=1, value=seg_name).font = DATA_FONT_BOLD_RED
        ws.cell(row=row, column=2, value=item["flag"]).font = DATA_FONT_BOLD_RED
        c3 = ws.cell(row=row, column=3, value=op_cur)
        c3.number_format = NUM_FMT_CURRENCY
        c3.font = DATA_FONT_RED if op_cur < 0 else DATA_FONT
        c4 = ws.cell(row=row, column=4, value=op_prev)
        c4.number_format = NUM_FMT_CURRENCY
        c5 = ws.cell(row=row, column=5, value=op_change)
        c5.number_format = NUM_FMT_CURRENCY
        c5.font = DATA_FONT_RED if op_change < 0 else DATA_FONT
        ws.cell(row=row, column=6, value=item["detail"]).font = SMALL_FONT

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            if col in [3, 4, 5]:
                cell.alignment = RIGHT_ALIGN
        ws.cell(row=row, column=6).alignment = LEFT_ALIGN

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 50

    # Freeze panes
    ws.freeze_panes = "A10"

    return ws


def build_sheet2_segment_analysis(wb, segment_data):
    """Build Sheet 2: セグメント分析"""
    ws = wb.create_sheet("セグメント分析")

    # Title
    ws.merge_cells("A1:K1")
    cell = ws.cell(row=1, column=1, value="セグメント別 P&L分析 202601")
    cell.font = TITLE_FONT
    cell.alignment = LEFT_ALIGN

    # Headers
    headers = ["セグメント", "P&L項目", "当月(202601)", "前月(202512)", "前月比", "前月比%",
               "3M平均", "3M乖離%", "6M平均", "6M乖離%", "フラグ"]
    for i, h in enumerate(headers, 1):
        apply_header_style(ws, 3, 1, len(headers))
        ws.cell(row=3, column=i, value=h)

    # Sort segments by absolute MoM operating profit impact
    def get_op_impact(seg_name):
        seg_d = segment_data.get(seg_name, {})
        op = seg_d.get("営業利益", {})
        return abs(op.get("mom_change", 0)) if op else 0

    # Filter out empty segments
    active_segments = [s for s in segment_data.keys()
                       if any(segment_data[s].get(item, {}).get("val_202601", 0) != 0
                              for item in PL_ITEMS_MAIN)]
    # Also include segments with non-zero prev month
    for s in segment_data.keys():
        if s not in active_segments:
            for item in PL_ITEMS_MAIN:
                d = segment_data.get(s, {}).get(item, {})
                if d and d.get("mom_change", 0) != 0:
                    active_segments.append(s)
                    break

    active_segments = list(set(active_segments))
    active_segments.sort(key=get_op_impact, reverse=True)

    pl_items = ["売上高", "変動費計", "限界利益", "人件費計", "固定費計", "営業利益"]

    current_row = 4
    seg_idx = 0
    for seg_name in active_segments:
        seg_d = segment_data.get(seg_name, {})
        if not seg_d:
            continue

        for pi_idx, pl_item in enumerate(pl_items):
            item_data = seg_d.get(pl_item, {})
            if not item_data:
                continue

            is_alt = seg_idx % 2 == 1

            # Only show segment name on first row of group
            if pi_idx == 0:
                ws.cell(row=current_row, column=1, value=seg_name).font = Font(name="Yu Gothic UI", size=10, bold=True)
            else:
                ws.cell(row=current_row, column=1, value="").font = DATA_FONT

            ws.cell(row=current_row, column=2, value=pl_item).font = DATA_FONT

            val = item_data.get("val_202601", 0)
            prev_val = val - item_data.get("mom_change", 0)
            mom_change = item_data.get("mom_change", 0)
            mom_pct = item_data.get("mom_pct", 0)
            avg_3m = item_data.get("avg_3m", 0)
            pct_3m = item_data.get("pct_3m", 0)
            avg_6m = item_data.get("avg_6m", 0)
            pct_6m = item_data.get("pct_6m", 0)
            flags = item_data.get("flags", "")

            c3 = ws.cell(row=current_row, column=3, value=val)
            c3.number_format = NUM_FMT_CURRENCY
            c4 = ws.cell(row=current_row, column=4, value=prev_val)
            c4.number_format = NUM_FMT_CURRENCY
            c5 = ws.cell(row=current_row, column=5, value=mom_change)
            c5.number_format = NUM_FMT_CURRENCY

            c6 = ws.cell(row=current_row, column=6, value=mom_pct / 100 if mom_pct != 0 else 0)
            c6.number_format = '+0.0%;-0.0%'

            c7 = ws.cell(row=current_row, column=7, value=avg_3m)
            c7.number_format = NUM_FMT_CURRENCY

            c8 = ws.cell(row=current_row, column=8, value=pct_3m / 100 if pct_3m != 0 else 0)
            c8.number_format = '+0.0%;-0.0%'

            c9 = ws.cell(row=current_row, column=9, value=avg_6m)
            c9.number_format = NUM_FMT_CURRENCY

            c10 = ws.cell(row=current_row, column=10, value=pct_6m / 100 if pct_6m != 0 else 0)
            c10.number_format = '+0.0%;-0.0%'

            ws.cell(row=current_row, column=11, value=flags).font = SMALL_FONT

            # Apply styles
            for col in range(1, 12):
                cell = ws.cell(row=current_row, column=col)
                cell.border = THIN_BORDER
                cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
                if col >= 3 and col <= 10:
                    cell.alignment = RIGHT_ALIGN

            ws.cell(row=current_row, column=11).alignment = LEFT_ALIGN

            # Conditional formatting for MoM%
            if abs(mom_pct) > 20:
                if mom_pct < -20:
                    ws.cell(row=current_row, column=6).fill = RED_FILL
                    ws.cell(row=current_row, column=6).font = DATA_FONT_RED
                elif mom_pct > 20:
                    ws.cell(row=current_row, column=6).fill = GREEN_FILL
                    ws.cell(row=current_row, column=6).font = DATA_FONT_GREEN

            # Bold red for flagged
            if flags:
                ws.cell(row=current_row, column=11).font = Font(name="Yu Gothic UI", size=9, color="FF0000")

            # Red font for negative values
            if val < 0:
                ws.cell(row=current_row, column=3).font = DATA_FONT_RED
            if mom_change < 0:
                ws.cell(row=current_row, column=5).font = DATA_FONT_RED

            current_row += 1

        seg_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 45

    ws.freeze_panes = "A4"

    return ws


def build_sheet3_trends(wb, analysis_data, segment_data):
    """Build Sheet 3: トレンド推移"""
    ws = wb.create_sheet("トレンド推移")

    # Title
    ws.merge_cells("A1:I1")
    cell = ws.cell(row=1, column=1, value="主要指標 8ヶ月トレンド 202506-202601")
    cell.font = TITLE_FONT
    cell.alignment = LEFT_ALIGN

    # --- Section 1: 全体P&L推移 ---
    write_section_header(ws, 3, 1, "■ 全体P&L推移", 9)

    trend_headers_1 = ["月"] + MONTH_LABELS
    for i, h in enumerate(trend_headers_1):
        col = i + 1
        apply_header_style(ws, 4, 1, len(trend_headers_1))
        ws.cell(row=4, column=col, value=h)

    trend_items = ["売上高", "変動費計", "限界利益", "固定費計", "営業利益"]
    trends = analysis_data["trends"]

    for idx, item in enumerate(trend_items):
        row = 5 + idx
        is_alt = idx % 2 == 1
        ws.cell(row=row, column=1, value=item).font = DATA_FONT

        vals = trends.get(item, [0]*8)
        for m_idx, val in enumerate(vals):
            c = ws.cell(row=row, column=2 + m_idx, value=val)
            c.number_format = NUM_FMT_CURRENCY
            c.font = DATA_FONT_RED if val < 0 else DATA_FONT

        for col in range(1, len(trend_headers_1) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            if col >= 2:
                cell.alignment = RIGHT_ALIGN

    # --- Section 2: セグメント別売上推移 (上位10) ---
    sec2_start = 12
    write_section_header(ws, sec2_start, 1, "■ セグメント別売上推移 (上位10)", 12)

    # Get top 10 segments by current revenue
    seg_revenue = {}
    for seg_name, seg_d in segment_data.items():
        rev_data = seg_d.get("売上高", {})
        if rev_data and rev_data.get("val_202601", 0) > 0:
            seg_revenue[seg_name] = rev_data.get("val_202601", 0)

    top10_rev = sorted(seg_revenue.items(), key=lambda x: x[1], reverse=True)[:10]
    top10_rev_names = [s[0] for s in top10_rev]

    # Headers
    rev_headers = ["月"] + top10_rev_names
    for i, h in enumerate(rev_headers):
        col = i + 1
        c = ws.cell(row=sec2_start + 1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

    # Data rows
    for m_idx in range(8):
        row = sec2_start + 2 + m_idx
        is_alt = m_idx % 2 == 1
        ws.cell(row=row, column=1, value=MONTH_LABELS[m_idx]).font = DATA_FONT

        for s_idx, seg_name in enumerate(top10_rev_names):
            seg_d = segment_data.get(seg_name, {})
            rev_data = seg_d.get("売上高", {})
            trend_vals = rev_data.get("trends", [0]*8)
            val = trend_vals[m_idx] if m_idx < len(trend_vals) else 0

            c = ws.cell(row=row, column=2 + s_idx, value=val)
            c.number_format = NUM_FMT_CURRENCY

        for col in range(1, len(rev_headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            if col >= 2:
                cell.alignment = RIGHT_ALIGN
                cell.font = DATA_FONT

    # --- Section 3: セグメント別営業利益推移 (上位10) ---
    sec3_start = sec2_start + 12
    write_section_header(ws, sec3_start, 1, "■ セグメント別営業利益推移 (上位10)", 12)

    # Get top 10 segments by absolute current OP
    seg_op = {}
    for seg_name, seg_d in segment_data.items():
        op_data = seg_d.get("営業利益", {})
        if op_data:
            seg_op[seg_name] = op_data.get("val_202601", 0)

    top10_op = sorted(seg_op.items(), key=lambda x: abs(x[1]), reverse=True)[:10]
    top10_op_names = [s[0] for s in top10_op]

    # Headers
    op_headers = ["月"] + top10_op_names
    for i, h in enumerate(op_headers):
        col = i + 1
        c = ws.cell(row=sec3_start + 1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

    # Data rows
    for m_idx in range(8):
        row = sec3_start + 2 + m_idx
        is_alt = m_idx % 2 == 1
        ws.cell(row=row, column=1, value=MONTH_LABELS[m_idx]).font = DATA_FONT

        for s_idx, seg_name in enumerate(top10_op_names):
            seg_d = segment_data.get(seg_name, {})
            op_data = seg_d.get("営業利益", {})
            trend_vals = op_data.get("trends", [0]*8)
            val = trend_vals[m_idx] if m_idx < len(trend_vals) else 0

            c = ws.cell(row=row, column=2 + s_idx, value=val)
            c.number_format = NUM_FMT_CURRENCY
            if val < 0:
                c.font = DATA_FONT_RED

        for col in range(1, len(op_headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
            if col >= 2:
                cell.alignment = RIGHT_ALIGN

    # Column widths
    ws.column_dimensions["A"].width = 10
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.freeze_panes = "B5"

    return ws


def build_sheet4_raw_data(wb, analysis_data):
    """Build Sheet 4: 全行詳細データ"""
    ws = wb.create_sheet("全行詳細データ")

    # Title
    ws.merge_cells("A1:G1")
    cell = ws.cell(row=1, column=1, value="202601 全P&Lデータ + 前月比")
    cell.font = TITLE_FONT
    cell.alignment = LEFT_ALIGN

    # Read the raw 202601 xlsx to get all rows
    filepath_202601 = os.path.join(CSV_DIR, "デジコン収益分析_202601.xlsx")
    df_202601 = load_xlsx_data(filepath_202601)

    filepath_202512 = os.path.join(CSV_DIR, "デジコン収益分析_202512.xlsx")
    df_202512 = load_xlsx_data(filepath_202512)

    # Find headers and structure
    header_row_idx = find_header_row(df_202601)

    # Write raw data from 202601
    # First write all rows
    start_row = 3

    # Get number of columns
    num_cols = df_202601.shape[1]

    # Determine the label for columns header
    # Write header row first
    ws.cell(row=start_row, column=1, value="行番号").font = HEADER_FONT

    # Write the actual xlsx headers
    if header_row_idx is not None:
        for col_idx in range(num_cols):
            val = df_202601.iloc[header_row_idx, col_idx]
            val_str = str(val) if pd.notna(val) else ""
            c = ws.cell(row=start_row, column=col_idx + 2, value=val_str)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER_ALIGN
            c.border = THIN_BORDER

    # Add comparison columns
    comp_start_col = num_cols + 2
    ws.cell(row=start_row, column=comp_start_col, value="202512合計").font = HEADER_FONT
    ws.cell(row=start_row, column=comp_start_col + 1, value="前月比").font = HEADER_FONT
    ws.cell(row=start_row, column=comp_start_col + 2, value="前月比%").font = HEADER_FONT

    for col in [comp_start_col, comp_start_col + 1, comp_start_col + 2]:
        c = ws.cell(row=start_row, column=col)
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.cell(row=start_row, column=1).border = THIN_BORDER
    ws.cell(row=start_row, column=1).alignment = CENTER_ALIGN

    # Write all data rows
    # Skip the first few metadata rows, start from a reasonable point
    data_start = 0
    for idx in range(df_202601.shape[0]):
        row_vals = df_202601.iloc[idx]
        # Check if this row has meaningful label content
        has_label = False
        for lc in [1, 2, 3, 4]:
            if lc < len(row_vals) and pd.notna(row_vals.iloc[lc]):
                val_str = str(row_vals.iloc[lc]).strip()
                if val_str and val_str != "nan":
                    has_label = True
                    break

        if idx == 0:
            # Always include first row (title etc.)
            pass

        excel_row = start_row + 1 + idx - data_start
        is_alt = (idx - data_start) % 2 == 1

        # Row number
        ws.cell(row=excel_row, column=1, value=idx + 1).font = DATA_FONT
        ws.cell(row=excel_row, column=1).border = THIN_BORDER
        ws.cell(row=excel_row, column=1).fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        for col_idx in range(num_cols):
            val = row_vals.iloc[col_idx]
            c = ws.cell(row=excel_row, column=col_idx + 2)

            if pd.notna(val):
                try:
                    float_val = float(val)
                    c.value = float_val
                    c.number_format = NUM_FMT_CURRENCY
                    if float_val < 0:
                        c.font = DATA_FONT_RED
                    else:
                        c.font = DATA_FONT
                except (ValueError, TypeError):
                    c.value = str(val)
                    c.font = DATA_FONT

            c.border = THIN_BORDER
            c.fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        # Add 202512 comparison for the 合計 column
        # Find the label for this row
        row_label = None
        for lc in [1, 2, 3, 4]:
            if lc < len(row_vals) and pd.notna(row_vals.iloc[lc]):
                lbl = str(row_vals.iloc[lc]).strip()
                if lbl and lbl != "nan":
                    row_label = lbl
                    break

        if row_label and idx < df_202512.shape[0]:
            # Try to find matching row in 202512
            val_202601_total = extract_totals_column(df_202601, idx)

            # Find matching row in 202512 by label
            matching_row_512 = find_row_by_label(df_202512, row_label)
            if matching_row_512 is not None:
                val_202512_total = extract_totals_column(df_202512, matching_row_512)
            else:
                val_202512_total = None

            if val_202512_total is not None and val_202601_total != 0:
                c_prev = ws.cell(row=excel_row, column=comp_start_col, value=val_202512_total)
                c_prev.number_format = NUM_FMT_CURRENCY
                c_prev.border = THIN_BORDER
                c_prev.fill = ALT_ROW_FILL if is_alt else WHITE_FILL

                mom = val_202601_total - val_202512_total
                c_mom = ws.cell(row=excel_row, column=comp_start_col + 1, value=mom)
                c_mom.number_format = NUM_FMT_CURRENCY
                c_mom.border = THIN_BORDER
                c_mom.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
                if mom < 0:
                    c_mom.font = DATA_FONT_RED

                if val_202512_total != 0:
                    mom_pct = mom / abs(val_202512_total)
                    c_pct = ws.cell(row=excel_row, column=comp_start_col + 2, value=mom_pct)
                    c_pct.number_format = '+0.0%;-0.0%'
                    c_pct.border = THIN_BORDER
                    c_pct.fill = ALT_ROW_FILL if is_alt else WHITE_FILL
                    apply_conditional_fill(c_pct, mom_pct * 100)

    # Column widths
    ws.column_dimensions["A"].width = 8
    for i in range(2, num_cols + 5):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.freeze_panes = "A4"

    return ws


# ===== Main =====

def main():
    print("Loading analysis data...")
    analysis_data = parse_analysis_csv()
    print(f"  Rankings: {len(analysis_data['rankings'])} items")
    print(f"  Trends: {len(analysis_data['trends'])} items")
    print(f"  Flagged: {len(analysis_data['flagged_segments'])} flags")

    print("Loading segment data...")
    segment_data = parse_segment_csv()
    print(f"  Segments: {len(segment_data)} segments")

    print("Creating workbook...")
    wb = Workbook()

    print("Building Sheet 1: エグゼクティブサマリー...")
    build_sheet1_executive_summary(wb, analysis_data, segment_data)

    print("Building Sheet 2: セグメント分析...")
    build_sheet2_segment_analysis(wb, segment_data)

    print("Building Sheet 3: トレンド推移...")
    build_sheet3_trends(wb, analysis_data, segment_data)

    print("Building Sheet 4: 全行詳細データ...")
    build_sheet4_raw_data(wb, analysis_data)

    print(f"Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done!")

    # Verify
    from openpyxl import load_workbook
    wb_check = load_workbook(OUTPUT_FILE)
    print(f"\nVerification:")
    for sheet_name in wb_check.sheetnames:
        ws = wb_check[sheet_name]
        print(f"  {sheet_name}: {ws.max_row} rows x {ws.max_column} cols")

    print(f"\nFile size: {os.path.getsize(OUTPUT_FILE):,} bytes")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
