#!/usr/bin/env python3
"""
デジコン収益分析 - Deep Segment-Level Analysis
Analyzes 8 months (202506-202601) of P&L data per segment.
"""

import openpyxl
import os
import csv
import sys
from collections import OrderedDict

BASE_DIR = '/Users/masaaki/Desktop/prm/report/boradmtg/csv/'
OUTPUT_CSV = '/Users/masaaki/Desktop/prm/report/boradmtg/segment_analysis_202601.csv'

# File configuration: period -> (filename, sheet_name)
FILES = OrderedDict([
    ('202506', ('デジコン収益分析_202506.xlsx', '2506月サマリー表 (2)')),
    ('202507', ('デジコン収益分析_202507.xlsx', '2507月サマリー表')),
    ('202508', ('デジコン収益分析_202508.xlsx', '2508月サマリー表')),
    ('202509', ('デジコン収益分析_202509.xlsx', '2509月サマリー表')),
    ('202510', ('デジコン収益分析_202510.xlsx', '2510月サマリー表')),
    ('202511', ('デジコン収益分析_202511.xlsx', '2511月サマリー表')),
    ('202512', ('デジコン収益分析_202512.xlsx', '2512月サマリー表 (2)')),
    ('202601', ('デジコン収益分析_202601.xlsx', '2601月サマリー表')),
])

PERIODS = list(FILES.keys())

# P&L items to extract (searched in columns 1-6 of each row)
PL_ITEMS = ['売上高', '変動費計', '限界利益', '人件費計', '固定費計', '営業利益']

# Segments to skip (subtotals and totals)
SKIP_SEGMENTS = {'小計', '合計'}


def find_key_rows(ws):
    """Find row numbers for each P&L item by scanning label columns."""
    key_rows = {}
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, 7):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                v_str = str(v).strip()
                for item in PL_ITEMS:
                    if v_str == item and item not in key_rows:
                        key_rows[item] = row_idx
    return key_rows


def get_segment_columns(ws):
    """Get segment name -> column index mapping from header row 4."""
    segments = OrderedDict()
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=col_idx).value
        if v is not None and col_idx >= 6:
            name = str(v).strip()
            if name not in SKIP_SEGMENTS:
                segments[name] = col_idx
    return segments


def extract_data(ws, key_rows, segment_cols):
    """Extract P&L values for each segment."""
    data = {}
    for seg_name, col_idx in segment_cols.items():
        seg_data = {}
        for item, row_idx in key_rows.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                val = 0
            try:
                val = float(val)
            except (ValueError, TypeError):
                val = 0
            seg_data[item] = val
        data[seg_name] = seg_data
    return data


def normalize_segment_name(name):
    """Normalize segment names for cross-file matching."""
    # Handle the 大占館 vs 大百科 difference in 202506
    if '大占館' in name:
        return name.replace('大占館', '大百科')
    return name


def load_all_data():
    """Load data from all 8 files."""
    all_data = {}  # {period: {segment_name: {pl_item: value}}}
    canonical_segments = None  # Use 202601 as canonical

    for period, (fname, sheet_name) in FILES.items():
        filepath = os.path.join(BASE_DIR, fname)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name]

        key_rows = find_key_rows(ws)
        segment_cols = get_segment_columns(ws)

        # Verify we found all P&L items
        missing = [item for item in PL_ITEMS if item not in key_rows]
        if missing:
            print(f"WARNING: {period} missing rows for: {missing}")

        data = extract_data(ws, key_rows, segment_cols)

        # Normalize segment names
        normalized_data = {}
        for seg_name, seg_vals in data.items():
            norm_name = normalize_segment_name(seg_name)
            normalized_data[norm_name] = seg_vals

        all_data[period] = normalized_data

        if period == '202601':
            canonical_segments = list(normalized_data.keys())

        wb.close()

    return all_data, canonical_segments


def fmt_k(val):
    """Format value in thousands (千円) with comma separator."""
    if val is None or val == 0:
        return "0"
    return f"{val:,.0f}"


def fmt_pct(val):
    """Format percentage."""
    if val is None:
        return "N/A"
    return f"{val:+.1f}%"


def safe_pct_change(new, old):
    """Calculate percentage change safely."""
    if old == 0:
        if new == 0:
            return 0.0
        return None  # undefined
    return ((new - old) / abs(old)) * 100


def main():
    print("=" * 100)
    print("  デジコン収益分析 - Deep Segment-Level Analysis (202506-202601)")
    print("=" * 100)
    print()

    # Load all data
    all_data, canonical_segments = load_all_data()
    print(f"Loaded {len(all_data)} periods, {len(canonical_segments)} segments")
    print(f"Segments: {', '.join(canonical_segments)}")
    print()

    # =========================================================================
    # Build analysis structures
    # =========================================================================

    # For each segment, compute metrics
    segment_analysis = {}

    for seg in canonical_segments:
        analysis = {}

        # 8-month trend for each P&L item
        for item in PL_ITEMS:
            trend = []
            for period in PERIODS:
                val = all_data.get(period, {}).get(seg, {}).get(item, 0)
                trend.append(val)
            analysis[f'{item}_trend'] = trend

            # Current value (202601)
            val_202601 = trend[7]  # index 7 = 202601
            analysis[f'{item}_202601'] = val_202601

            # MoM change vs 202512
            val_202512 = trend[6]  # index 6 = 202512
            mom_abs = val_202601 - val_202512
            mom_pct = safe_pct_change(val_202601, val_202512)
            analysis[f'{item}_mom_abs'] = mom_abs
            analysis[f'{item}_mom_pct'] = mom_pct

            # 3-month average (202510-202512) = indices 4,5,6
            avg_3m = sum(trend[4:7]) / 3 if len(trend) >= 7 else 0
            dev_3m = val_202601 - avg_3m
            dev_3m_pct = safe_pct_change(val_202601, avg_3m) if avg_3m != 0 else None
            analysis[f'{item}_avg3m'] = avg_3m
            analysis[f'{item}_dev3m_abs'] = dev_3m
            analysis[f'{item}_dev3m_pct'] = dev_3m_pct

            # 6-month average (202507-202512) = indices 1,2,3,4,5,6
            avg_6m = sum(trend[1:7]) / 6 if len(trend) >= 7 else 0
            dev_6m = val_202601 - avg_6m
            dev_6m_pct = safe_pct_change(val_202601, avg_6m) if avg_6m != 0 else None
            analysis[f'{item}_avg6m'] = avg_6m
            analysis[f'{item}_dev6m_abs'] = dev_6m
            analysis[f'{item}_dev6m_pct'] = dev_6m_pct

        segment_analysis[seg] = analysis

    # =========================================================================
    # FLAGS
    # =========================================================================

    flags = {}
    for seg in canonical_segments:
        a = segment_analysis[seg]
        seg_flags = []

        # Flag 1: Operating profit turned negative (黒字→赤字)
        op_202512 = a['営業利益_trend'][6]
        op_202601 = a['営業利益_202601']
        if op_202512 > 0 and op_202601 < 0:
            seg_flags.append(f"黒字→赤字転落: {fmt_k(op_202512)} → {fmt_k(op_202601)}")

        # Flag 2: Revenue declined >10% MoM
        rev_mom_pct = a['売上高_mom_pct']
        if rev_mom_pct is not None and rev_mom_pct < -10:
            seg_flags.append(f"売上高MoM {fmt_pct(rev_mom_pct)} (>10%減少)")

        # Flag 3: Cost grew faster than revenue (variable + fixed cost growth > revenue growth)
        rev_pct = a['売上高_mom_pct']
        var_cost_pct = a['変動費計_mom_pct']
        fixed_cost_pct = a['固定費計_mom_pct']
        if rev_pct is not None and var_cost_pct is not None:
            # Total cost change
            total_cost_202601 = a['変動費計_202601'] + a['固定費計_202601']
            total_cost_202512 = a['変動費計_trend'][6] + a['固定費計_trend'][6]
            cost_pct = safe_pct_change(total_cost_202601, total_cost_202512)
            if cost_pct is not None and rev_pct is not None:
                if cost_pct > rev_pct and a['売上高_202601'] > 0:
                    seg_flags.append(f"コスト増>売上増: 売上{fmt_pct(rev_pct)}, コスト{fmt_pct(cost_pct)}")

        # Flag 4: 3M/6M average deviation exceeds ±20%
        for item in ['売上高', '営業利益']:
            dev3m = a[f'{item}_dev3m_pct']
            dev6m = a[f'{item}_dev6m_pct']
            if dev3m is not None and abs(dev3m) > 20:
                seg_flags.append(f"{item} 3M平均乖離 {fmt_pct(dev3m)}")
            if dev6m is not None and abs(dev6m) > 20:
                seg_flags.append(f"{item} 6M平均乖離 {fmt_pct(dev6m)}")

        if seg_flags:
            flags[seg] = seg_flags

    # =========================================================================
    # RANKINGS
    # =========================================================================

    # Ranking by MoM operating profit change (descending absolute impact)
    op_mom_ranking = sorted(
        canonical_segments,
        key=lambda s: segment_analysis[s]['営業利益_mom_abs'],
        reverse=True
    )

    # Ranking by MoM revenue change (descending absolute impact)
    rev_mom_ranking = sorted(
        canonical_segments,
        key=lambda s: segment_analysis[s]['売上高_mom_abs'],
        reverse=True
    )

    # =========================================================================
    # PRINT SUMMARY TO STDOUT
    # =========================================================================

    # --- Section 1: Top 10 by MoM Operating Profit Impact ---
    print("=" * 100)
    print("  [1] Segment Ranking by MoM Operating Profit Change (202512→202601)")
    print("=" * 100)
    print(f"{'Rank':<5} {'Segment':<28} {'202601':>14} {'202512':>14} {'MoM Change':>14} {'MoM %':>10}")
    print("-" * 100)
    for i, seg in enumerate(op_mom_ranking[:10], 1):
        a = segment_analysis[seg]
        val601 = a['営業利益_202601']
        val512 = a['営業利益_trend'][6]
        mom = a['営業利益_mom_abs']
        pct = a['営業利益_mom_pct']
        pct_str = fmt_pct(pct) if pct is not None else "N/A"
        print(f"{i:<5} {seg:<28} {fmt_k(val601):>14} {fmt_k(val512):>14} {fmt_k(mom):>14} {pct_str:>10}")
    print()
    print("  ... Bottom 5 (worst declines):")
    print(f"{'Rank':<5} {'Segment':<28} {'202601':>14} {'202512':>14} {'MoM Change':>14} {'MoM %':>10}")
    print("-" * 100)
    for i, seg in enumerate(reversed(op_mom_ranking[-5:]), 1):
        a = segment_analysis[seg]
        val601 = a['営業利益_202601']
        val512 = a['営業利益_trend'][6]
        mom = a['営業利益_mom_abs']
        pct = a['営業利益_mom_pct']
        pct_str = fmt_pct(pct) if pct is not None else "N/A"
        print(f"  {i:<3} {seg:<28} {fmt_k(val601):>14} {fmt_k(val512):>14} {fmt_k(mom):>14} {pct_str:>10}")
    print()

    # --- Section 2: Top 10 by MoM Revenue Impact ---
    print("=" * 100)
    print("  [2] Segment Ranking by MoM Revenue Change (202512→202601)")
    print("=" * 100)
    print(f"{'Rank':<5} {'Segment':<28} {'202601':>14} {'202512':>14} {'MoM Change':>14} {'MoM %':>10}")
    print("-" * 100)
    for i, seg in enumerate(rev_mom_ranking[:10], 1):
        a = segment_analysis[seg]
        val601 = a['売上高_202601']
        val512 = a['売上高_trend'][6]
        mom = a['売上高_mom_abs']
        pct = a['売上高_mom_pct']
        pct_str = fmt_pct(pct) if pct is not None else "N/A"
        print(f"{i:<5} {seg:<28} {fmt_k(val601):>14} {fmt_k(val512):>14} {fmt_k(mom):>14} {pct_str:>10}")
    print()
    print("  ... Bottom 5 (worst declines):")
    print(f"{'Rank':<5} {'Segment':<28} {'202601':>14} {'202512':>14} {'MoM Change':>14} {'MoM %':>10}")
    print("-" * 100)
    for i, seg in enumerate(reversed(rev_mom_ranking[-5:]), 1):
        a = segment_analysis[seg]
        val601 = a['売上高_202601']
        val512 = a['売上高_trend'][6]
        mom = a['売上高_mom_abs']
        pct = a['売上高_mom_pct']
        pct_str = fmt_pct(pct) if pct is not None else "N/A"
        print(f"  {i:<3} {seg:<28} {fmt_k(val601):>14} {fmt_k(val512):>14} {fmt_k(mom):>14} {pct_str:>10}")
    print()

    # --- Section 3: Flagged Segments ---
    print("=" * 100)
    print("  [3] Flagged Segments (Issues Detected)")
    print("=" * 100)
    if flags:
        for seg, seg_flags in sorted(flags.items()):
            print(f"\n  ** {seg} **")
            for f in seg_flags:
                print(f"     - {f}")
    else:
        print("  No segments flagged.")
    print()

    # --- Section 4: 8-month Revenue Trend for Top 5 segments ---
    # Top 5 by absolute revenue
    top5_rev = sorted(canonical_segments, key=lambda s: segment_analysis[s]['売上高_202601'], reverse=True)[:5]
    print("=" * 100)
    print("  [4] 8-Month Revenue Trend (売上高) - Top 5 Segments by 202601 Revenue")
    print("=" * 100)
    header = f"{'Segment':<28}"
    for p in PERIODS:
        header += f" {p[2:]:>10}"
    print(header)
    print("-" * 100)
    for seg in top5_rev:
        a = segment_analysis[seg]
        trend = a['売上高_trend']
        line = f"{seg:<28}"
        for v in trend:
            line += f" {fmt_k(v):>10}"
        print(line)
    print()

    # --- Section 5: 8-month Operating Profit Trend for Top 5 segments ---
    top5_op = sorted(canonical_segments, key=lambda s: abs(segment_analysis[s]['営業利益_mom_abs']), reverse=True)[:5]
    print("=" * 100)
    print("  [5] 8-Month Operating Profit Trend (営業利益) - Top 5 by MoM Impact")
    print("=" * 100)
    header = f"{'Segment':<28}"
    for p in PERIODS:
        header += f" {p[2:]:>10}"
    print(header)
    print("-" * 100)
    for seg in top5_op:
        a = segment_analysis[seg]
        trend = a['営業利益_trend']
        line = f"{seg:<28}"
        for v in trend:
            line += f" {fmt_k(v):>10}"
        print(line)
    print()

    # --- Section 6: Per-Segment P&L Summary ---
    print("=" * 100)
    print("  [6] Per-Segment P&L Summary (All Segments)")
    print("=" * 100)
    for seg in canonical_segments:
        a = segment_analysis[seg]
        print(f"\n  ---- {seg} ----")
        print(f"  {'P&L Item':<16} {'202601':>14} {'MoM Chg':>14} {'MoM %':>10} {'3M Avg':>14} {'3M Dev':>14} {'3M Dev%':>10} {'6M Avg':>14} {'6M Dev':>14} {'6M Dev%':>10}")
        print(f"  {'-'*140}")
        for item in PL_ITEMS:
            val = a[f'{item}_202601']
            mom = a[f'{item}_mom_abs']
            mom_p = a[f'{item}_mom_pct']
            avg3 = a[f'{item}_avg3m']
            d3 = a[f'{item}_dev3m_abs']
            d3p = a[f'{item}_dev3m_pct']
            avg6 = a[f'{item}_avg6m']
            d6 = a[f'{item}_dev6m_abs']
            d6p = a[f'{item}_dev6m_pct']
            print(f"  {item:<16} {fmt_k(val):>14} {fmt_k(mom):>14} {fmt_pct(mom_p) if mom_p is not None else 'N/A':>10} "
                  f"{fmt_k(avg3):>14} {fmt_k(d3):>14} {fmt_pct(d3p) if d3p is not None else 'N/A':>10} "
                  f"{fmt_k(avg6):>14} {fmt_k(d6):>14} {fmt_pct(d6p) if d6p is not None else 'N/A':>10}")
        # Mini 8-month trend
        print(f"  Revenue trend:  ", end="")
        for i, p in enumerate(PERIODS):
            print(f" {p[2:]}={fmt_k(a['売上高_trend'][i])}", end="")
        print()
        print(f"  OP trend:       ", end="")
        for i, p in enumerate(PERIODS):
            print(f" {p[2:]}={fmt_k(a['営業利益_trend'][i])}", end="")
        print()

    print()

    # =========================================================================
    # SAVE DETAILED CSV
    # =========================================================================

    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)

        # Header
        headers = ['Segment', 'P&L Item',
                   '202601 Value', 'MoM Change (abs)', 'MoM Change (%)',
                   '3M Avg (202510-12)', '3M Deviation (abs)', '3M Deviation (%)',
                   '6M Avg (202507-12)', '6M Deviation (abs)', '6M Deviation (%)']
        for p in PERIODS:
            headers.append(f'Trend_{p}')
        headers.append('Flags')
        writer.writerow(headers)

        for seg in canonical_segments:
            a = segment_analysis[seg]
            seg_flag_str = ' | '.join(flags.get(seg, []))

            for item in PL_ITEMS:
                row = [
                    seg, item,
                    a[f'{item}_202601'],
                    a[f'{item}_mom_abs'],
                    a[f'{item}_mom_pct'] if a[f'{item}_mom_pct'] is not None else '',
                    a[f'{item}_avg3m'],
                    a[f'{item}_dev3m_abs'],
                    a[f'{item}_dev3m_pct'] if a[f'{item}_dev3m_pct'] is not None else '',
                    a[f'{item}_avg6m'],
                    a[f'{item}_dev6m_abs'],
                    a[f'{item}_dev6m_pct'] if a[f'{item}_dev6m_pct'] is not None else '',
                ]
                for v in a[f'{item}_trend']:
                    row.append(v)
                # Only add flags on first item row for this segment
                if item == PL_ITEMS[0]:
                    row.append(seg_flag_str)
                else:
                    row.append('')
                writer.writerow(row)

        # Add a ranking summary section
        writer.writerow([])
        writer.writerow(['=== MoM Operating Profit Ranking ==='])
        writer.writerow(['Rank', 'Segment', '202601 OP', '202512 OP', 'MoM Change', 'MoM %'])
        for i, seg in enumerate(op_mom_ranking, 1):
            a = segment_analysis[seg]
            writer.writerow([
                i, seg,
                a['営業利益_202601'],
                a['営業利益_trend'][6],
                a['営業利益_mom_abs'],
                a['営業利益_mom_pct'] if a['営業利益_mom_pct'] is not None else ''
            ])

        writer.writerow([])
        writer.writerow(['=== MoM Revenue Ranking ==='])
        writer.writerow(['Rank', 'Segment', '202601 Revenue', '202512 Revenue', 'MoM Change', 'MoM %'])
        for i, seg in enumerate(rev_mom_ranking, 1):
            a = segment_analysis[seg]
            writer.writerow([
                i, seg,
                a['売上高_202601'],
                a['売上高_trend'][6],
                a['売上高_mom_abs'],
                a['売上高_mom_pct'] if a['売上高_mom_pct'] is not None else ''
            ])

        writer.writerow([])
        writer.writerow(['=== Flagged Segments ==='])
        writer.writerow(['Segment', 'Flag'])
        for seg, seg_flags in sorted(flags.items()):
            for fl in seg_flags:
                writer.writerow([seg, fl])

    print(f"\nDetailed CSV saved to: {OUTPUT_CSV}")
    print("=" * 100)


if __name__ == '__main__':
    main()
