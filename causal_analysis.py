#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deep causal analysis: 売上高 (Revenue) vs 広告宣伝費 (Advertising Costs)
across 8 months (202506-202601), by segment and total.
"""
import openpyxl
import os
import math

BASE_DIR = "/Users/masaaki/Desktop/prm/report/boradmtg/csv"
MONTHS = ["202506", "202507", "202508", "202509", "202510", "202511", "202512", "202601"]
MONTH_LABELS = ["25/06", "25/07", "25/08", "25/09", "25/10", "25/11", "25/12", "26/01"]

# Segment columns mapping
# 202506-202509: segments at cols 6-27, 小計=28, プロモ=29, COIPO=30, その他=31, 合計=32
# 202510-202601: segments at cols 7-28, 小計=29, プロモ=30, COIPO=31, その他=32, 合計=33

# Segment names (from 202601 file, row 4)
SEGMENT_NAMES_NEW = {
    7: "モバイル新規", 8: "モバイル新規(27)_運用", 9: "リニューアル(27)_運用",
    10: "リニューアル(28)候補", 11: "新規_メディア(大百科)", 12: "よしもと大集合",
    13: "モバイル注力", 14: "はやとも", 15: "注力外", 16: "はやともヤースー",
    17: "橋本京明", 18: "めくる/ヤースー", 19: "ギャル", 20: "JUNO",
    21: "Love/sLove", 22: "ISP", 23: "プラットフォーム", 24: "アプリ",
    25: "海外展開", 26: "自社メディア", 27: "新規_ロト", 28: "新規_ポイント"
}

# For old format, cols shift by -1
SEGMENT_NAMES_OLD = {k-1: v for k, v in SEGMENT_NAMES_NEW.items()}

def find_summary_sheet(wb):
    for name in wb.sheetnames:
        if 'サマリー' in name:
            return wb[name]
    return wb[wb.sheetnames[0]]

def get_value(ws, row, col):
    """Get numeric value from cell, return 0 if None."""
    v = ws.cell(row, col).value
    if v is None:
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0

def find_row_by_label(ws, target_labels, search_cols=range(2,7)):
    """Find row number by searching for label text in given columns."""
    for r in range(1, ws.max_row+1):
        for c in search_cols:
            v = ws.cell(r, c).value
            if v is not None and isinstance(v, str):
                for label in target_labels:
                    if label == v.strip():
                        return r
    return None

def is_new_format(month):
    """202510 and later have 33-col format."""
    return int(month) >= 202510

def get_total_col(month):
    return 33 if is_new_format(month) else 32

def get_segment_cols(month):
    """Return dict of {col: segment_name}."""
    if is_new_format(month):
        return SEGMENT_NAMES_NEW
    else:
        return SEGMENT_NAMES_OLD

# ==================== DATA EXTRACTION ====================
print("=" * 100)
print("データ抽出中...")
print("=" * 100)

# Storage: {month: {metric: {segment_name: value, "合計": value}}}
data = {}

for month in MONTHS:
    fpath = os.path.join(BASE_DIR, f"デジコン収益分析_{month}.xlsx")
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = find_summary_sheet(wb)

    total_col = get_total_col(month)
    seg_cols = get_segment_cols(month)

    # Find key rows
    row_revenue = find_row_by_label(ws, ["売上高"])
    row_ad = find_row_by_label(ws, ["広告宣伝費"])
    row_var_total = find_row_by_label(ws, ["変動費計"])
    row_marginal = find_row_by_label(ws, ["限界利益"])
    row_op_profit = find_row_by_label(ws, ["営業利益"])

    month_data = {}

    # Extract revenue by segment
    rev_seg = {}
    ad_seg = {}
    for col, name in seg_cols.items():
        rev_seg[name] = get_value(ws, row_revenue, col)
        ad_seg[name] = get_value(ws, row_ad, col)

    rev_seg["合計"] = get_value(ws, row_revenue, total_col)
    ad_seg["合計"] = get_value(ws, row_ad, total_col)

    month_data["売上高"] = rev_seg
    month_data["広告宣伝費"] = ad_seg
    month_data["変動費計"] = get_value(ws, row_var_total, total_col)
    month_data["限界利益"] = get_value(ws, row_marginal, total_col)
    month_data["営業利益"] = get_value(ws, row_op_profit, total_col)

    data[month] = month_data
    wb.close()

# Verify extracted data
print("\n検証: 各月の合計売上高・広告宣伝費")
for m in MONTHS:
    rev = data[m]["売上高"]["合計"]
    ad = data[m]["広告宣伝費"]["合計"]
    print(f"  {m}: 売上高={rev:,.0f}  広告宣伝費={ad:,.0f}")

# ==================== HELPER FUNCTIONS ====================
def fmt(v, w=14):
    """Format number with commas, right-aligned."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "-".rjust(w)
    return f"{v:,.0f}".rjust(w)

def fmt_pct(v, w=10):
    """Format percentage."""
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return "-".rjust(w)
    return f"{v:.1f}%".rjust(w)

def fmt_r(v, w=10):
    """Format correlation coefficient."""
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return "-".rjust(w)
    return f"{v:.4f}".rjust(w)

def calc_corr(x, y):
    """Calculate Pearson correlation between two lists."""
    n = len(x)
    if n < 3:
        return None
    mx = sum(x) / n
    my = sum(y) / n
    sx = sum((xi - mx)**2 for xi in x)
    sy = sum((yi - my)**2 for yi in y)
    sxy = sum((xi - mx) * (yi - my) for xi, yi in zip(x, y))
    if sx == 0 or sy == 0:
        return None
    return sxy / (sx**0.5 * sy**0.5)

def mom_change(vals):
    """Calculate month-over-month changes."""
    return [vals[i] - vals[i-1] for i in range(1, len(vals))]

def mom_pct_change(vals):
    """Calculate MoM % changes."""
    result = []
    for i in range(1, len(vals)):
        if vals[i-1] != 0:
            result.append((vals[i] - vals[i-1]) / vals[i-1] * 100)
        else:
            result.append(None)
    return result

# ==================== ANALYSIS 1: 全体推移 ====================
print("\n")
print("=" * 100)
print("【分析1】全体: 売上高 vs 広告宣伝費の月次推移")
print("=" * 100)

revenues = [data[m]["売上高"]["合計"] for m in MONTHS]
ad_costs = [data[m]["広告宣伝費"]["合計"] for m in MONTHS]
var_totals = [data[m]["変動費計"] for m in MONTHS]
marginals = [data[m]["限界利益"] for m in MONTHS]
op_profits = [data[m]["営業利益"] for m in MONTHS]

# Table 1a: Absolute values
print("\n■ 月次絶対値")
header = "月".ljust(8) + "売上高".rjust(16) + "広告宣伝費".rjust(16) + "広告費/売上比率".rjust(14) + "変動費計".rjust(16) + "限界利益".rjust(16) + "営業利益".rjust(16)
print(header)
print("-" * len(header))
for i, m in enumerate(MONTHS):
    ratio = ad_costs[i] / revenues[i] * 100 if revenues[i] != 0 else 0
    print(f"{MONTH_LABELS[i].ljust(8)}{fmt(revenues[i], 16)}{fmt(ad_costs[i], 16)}{fmt_pct(ratio, 14)}{fmt(var_totals[i], 16)}{fmt(marginals[i], 16)}{fmt(op_profits[i], 16)}")

# Table 1b: MoM changes
print("\n■ 前月比変動（MoM）")
rev_chg = mom_change(revenues)
ad_chg = mom_change(ad_costs)
rev_pct = mom_pct_change(revenues)
ad_pct = mom_pct_change(ad_costs)

header2 = "月".ljust(8) + "Δ売上高".rjust(16) + "Δ売上(%)".rjust(10) + "Δ広告宣伝費".rjust(16) + "Δ広告(%)".rjust(10) + "方向一致".rjust(10)
print(header2)
print("-" * len(header2))
for i in range(len(rev_chg)):
    same_dir = "○" if (rev_chg[i] > 0 and ad_chg[i] > 0) or (rev_chg[i] < 0 and ad_chg[i] < 0) else "×"
    print(f"{MONTH_LABELS[i+1].ljust(8)}{fmt(rev_chg[i], 16)}{fmt_pct(rev_pct[i], 10)}{fmt(ad_chg[i], 16)}{fmt_pct(ad_pct[i], 10)}{same_dir.rjust(10)}")

# ==================== ANALYSIS 2: タイムラグ分析 ====================
print("\n")
print("=" * 100)
print("【分析2】タイムラグ分析（Lagged Correlation）")
print("=" * 100)

# Level correlations
print("\n■ 水準値の相関分析")

# Same month: ad(t) vs rev(t)
corr_same = calc_corr(ad_costs, revenues)
# Lag 1: ad(t) vs rev(t+1)
corr_lag1 = calc_corr(ad_costs[:-1], revenues[1:])
# Reverse: rev(t-1) vs ad(t) -> ad(t) follows rev(t-1)?
corr_rev = calc_corr(revenues[:-1], ad_costs[1:])

print(f"  広告宣伝費(t) vs 売上高(t)   [同月相関]:       r = {fmt_r(corr_same)}")
print(f"  広告宣伝費(t) vs 売上高(t+1) [1ヶ月ラグ]:      r = {fmt_r(corr_lag1)}")
print(f"  売上高(t) vs 広告宣伝費(t+1) [逆因果チェック]:  r = {fmt_r(corr_rev)}")

# Delta correlations (MoM changes)
print("\n■ 変動値（前月差分）の相関分析")
rev_delta = mom_change(revenues)
ad_delta = mom_change(ad_costs)

# Same month delta
corr_d_same = calc_corr(ad_delta, rev_delta)
# Lag 1 delta: Δad(t) vs Δrev(t+1)
corr_d_lag1 = calc_corr(ad_delta[:-1], rev_delta[1:])
# Reverse delta
corr_d_rev = calc_corr(rev_delta[:-1], ad_delta[1:])

print(f"  Δ広告宣伝費(t) vs Δ売上高(t)   [同月変動相関]:       r = {fmt_r(corr_d_same)}")
print(f"  Δ広告宣伝費(t) vs Δ売上高(t+1) [1ヶ月ラグ変動]:      r = {fmt_r(corr_d_lag1)}")
print(f"  Δ売上高(t) vs Δ広告宣伝費(t+1) [逆因果変動チェック]:  r = {fmt_r(corr_d_rev)}")

print("\n■ 解釈ガイド")
print("  r > 0.7: 強い正の相関  |  0.3 < r < 0.7: 中程度  |  r < 0.3: 弱い")
print("  ラグ相関 > 同月相関: 広告効果に遅延あり")
print("  逆因果相関が高い: 売上に応じて広告予算が調整されている可能性")

# ==================== ANALYSIS 3: セグメント別分析 ====================
print("\n")
print("=" * 100)
print("【分析3】セグメント別: 広告宣伝費の配分と売上の関係")
print("=" * 100)

# Get all segment names (use new format names)
all_segments = list(SEGMENT_NAMES_NEW.values())

# For each segment, get 8-month revenue and ad costs
seg_data = {}
for seg in all_segments:
    seg_revs = []
    seg_ads = []
    for m in MONTHS:
        seg_revs.append(data[m]["売上高"].get(seg, 0))
        seg_ads.append(data[m]["広告宣伝費"].get(seg, 0))
    seg_data[seg] = {"rev": seg_revs, "ad": seg_ads}

# Filter segments with non-zero ad spend
active_segments = []
for seg in all_segments:
    total_ad = sum(seg_data[seg]["ad"])
    if total_ad > 0:
        active_segments.append(seg)

print(f"\n広告宣伝費が発生しているセグメント数: {len(active_segments)} / {len(all_segments)}")

for seg in active_segments:
    revs = seg_data[seg]["rev"]
    ads = seg_data[seg]["ad"]

    print(f"\n{'─'*80}")
    print(f"■ {seg}")
    print(f"{'─'*80}")

    header3 = "月".ljust(8) + "売上高".rjust(14) + "広告宣伝費".rjust(14) + "広告/売上比率".rjust(14)
    print(header3)
    for i, m in enumerate(MONTHS):
        ratio = ads[i] / revs[i] * 100 if revs[i] != 0 else 0
        print(f"{MONTH_LABELS[i].ljust(8)}{fmt(revs[i], 14)}{fmt(ads[i], 14)}{fmt_pct(ratio, 14)}")

    # MoM correlation for this segment
    if len(revs) >= 3:
        seg_rev_delta = mom_change(revs)
        seg_ad_delta = mom_change(ads)
        seg_corr = calc_corr(seg_ad_delta, seg_rev_delta)

        # Check: ad increased but rev didn't
        flags = []
        for i in range(len(seg_ad_delta)):
            if seg_ad_delta[i] > 0 and seg_rev_delta[i] <= 0:
                flags.append(MONTH_LABELS[i+1])

        print(f"  MoM変動相関: r = {fmt_r(seg_corr)}")
        if flags:
            print(f"  ⚠ 広告増→売上減の月: {', '.join(flags)}")

# ==================== ANALYSIS 4: 広告効率分析 ====================
print("\n")
print("=" * 100)
print("【分析4】広告効率分析")
print("=" * 100)

print("\n■ 月次広告効率指標")
header4 = "月".ljust(8) + "売上高/広告費".rjust(16) + "Δ売上/Δ広告費".rjust(16) + "限界効率判定".rjust(14)
print(header4)
print("-" * len(header4))

rev_per_ad = []
marginal_eff = []
for i, m in enumerate(MONTHS):
    rpa = revenues[i] / ad_costs[i] if ad_costs[i] != 0 else None
    rev_per_ad.append(rpa)

    if i == 0:
        me = None
    else:
        d_ad = ad_costs[i] - ad_costs[i-1]
        d_rev = revenues[i] - revenues[i-1]
        if d_ad != 0:
            me = d_rev / d_ad
        else:
            me = None
    marginal_eff.append(me)

    rpa_str = f"{rpa:.2f}倍" if rpa is not None else "-"
    me_str = f"{me:.2f}倍" if me is not None else "-"

    if me is not None:
        if me > 1:
            judgment = "効率的"
        elif me > 0:
            judgment = "非効率"
        else:
            judgment = "逆効果"
    else:
        judgment = "-"

    print(f"{MONTH_LABELS[i].ljust(8)}{rpa_str.rjust(16)}{me_str.rjust(16)}{judgment.rjust(14)}")

print("\n■ 広告効率のトレンド")
valid_rpa = [(i, v) for i, v in enumerate(rev_per_ad) if v is not None]
if len(valid_rpa) >= 2:
    first_rpa = valid_rpa[0][1]
    last_rpa = valid_rpa[-1][1]
    trend = "改善" if last_rpa > first_rpa else "悪化"
    print(f"  広告1円あたり売上: {first_rpa:.2f}倍 ({MONTH_LABELS[valid_rpa[0][0]]}) → {last_rpa:.2f}倍 ({MONTH_LABELS[valid_rpa[-1][0]]}) → {trend}")

valid_me = [(i, v) for i, v in enumerate(marginal_eff) if v is not None]
pos_me = sum(1 for _, v in valid_me if v > 0)
neg_me = sum(1 for _, v in valid_me if v <= 0)
print(f"  限界効率: 正の月={pos_me}回, 負/ゼロの月={neg_me}回")

# ==================== ANALYSIS 5: 広告宣伝費の集中度 ====================
print("\n")
print("=" * 100)
print("【分析5】広告宣伝費の集中度分析")
print("=" * 100)

# Ad spend distribution for 202506 and 202601
for idx, (m, label) in enumerate([(MONTHS[0], MONTH_LABELS[0]), (MONTHS[-1], MONTH_LABELS[-1])]):
    print(f"\n■ {label} 広告宣伝費 セグメント別配分")
    total_ad = data[m]["広告宣伝費"]["合計"]

    seg_ad_list = []
    for seg in all_segments:
        ad_val = data[m]["広告宣伝費"].get(seg, 0)
        if ad_val > 0:
            seg_ad_list.append((seg, ad_val))

    seg_ad_list.sort(key=lambda x: x[1], reverse=True)

    header5 = "セグメント".ljust(30) + "広告宣伝費".rjust(14) + "構成比".rjust(10) + "累積構成比".rjust(10)
    print(header5)
    print("-" * len(header5))

    cum_pct = 0
    for seg, ad_val in seg_ad_list:
        pct = ad_val / total_ad * 100 if total_ad != 0 else 0
        cum_pct += pct
        print(f"{seg.ljust(30)}{fmt(ad_val, 14)}{fmt_pct(pct, 10)}{fmt_pct(cum_pct, 10)}")
    print(f"{'合計'.ljust(30)}{fmt(total_ad, 14)}{fmt_pct(100, 10)}")

# Distribution shift analysis
print(f"\n■ 構成比シフト分析 ({MONTH_LABELS[0]} → {MONTH_LABELS[-1]})")
header6 = "セグメント".ljust(30) + f"{MONTH_LABELS[0]}構成比".rjust(12) + f"{MONTH_LABELS[-1]}構成比".rjust(12) + "変動幅(pp)".rjust(12)
print(header6)
print("-" * len(header6))

total_ad_first = data[MONTHS[0]]["広告宣伝費"]["合計"]
total_ad_last = data[MONTHS[-1]]["広告宣伝費"]["合計"]

shift_data = []
for seg in all_segments:
    ad_first = data[MONTHS[0]]["広告宣伝費"].get(seg, 0)
    ad_last = data[MONTHS[-1]]["広告宣伝費"].get(seg, 0)
    pct_first = ad_first / total_ad_first * 100 if total_ad_first != 0 else 0
    pct_last = ad_last / total_ad_last * 100 if total_ad_last != 0 else 0
    shift = pct_last - pct_first
    if abs(ad_first) + abs(ad_last) > 0:
        shift_data.append((seg, pct_first, pct_last, shift))

shift_data.sort(key=lambda x: abs(x[3]), reverse=True)

for seg, pf, pl, sh in shift_data:
    sign = "+" if sh >= 0 else ""
    print(f"{seg.ljust(30)}{fmt_pct(pf, 12)}{fmt_pct(pl, 12)}{(sign + f'{sh:.1f}pp').rjust(12)}")

# ==================== ANALYSIS 6: 因果関係の結論 ====================
print("\n")
print("=" * 100)
print("【分析6】因果関係の結論")
print("=" * 100)

# a. 広告費を増やした結果、売上は増えたのか？
print("\n■ a. 広告費を増やした結果、売上は増えたのか？")
total_ad_change = ad_costs[-1] - ad_costs[0]
total_rev_change = revenues[-1] - revenues[0]
ad_change_pct = total_ad_change / ad_costs[0] * 100
rev_change_pct = total_rev_change / revenues[0] * 100

print(f"  期間全体（{MONTH_LABELS[0]}→{MONTH_LABELS[-1]}）:")
print(f"    広告宣伝費: {ad_costs[0]:,.0f} → {ad_costs[-1]:,.0f} ({ad_change_pct:+.1f}%)")
print(f"    売上高:     {revenues[0]:,.0f} → {revenues[-1]:,.0f} ({rev_change_pct:+.1f}%)")

# Count months where both moved same direction
same_dir_count = sum(1 for i in range(len(rev_chg)) if (rev_chg[i] > 0 and ad_chg[i] > 0) or (rev_chg[i] < 0 and ad_chg[i] < 0))
print(f"  MoM方向一致: {same_dir_count}/{len(rev_chg)}ヶ月")
print(f"  同月変動相関: r = {corr_d_same:.4f}" if corr_d_same else "  同月変動相関: 計算不可")

if total_ad_change > 0 and total_rev_change > 0:
    print("  → 広告費増加と売上増加が同時に発生")
elif total_ad_change > 0 and total_rev_change <= 0:
    print("  → 広告費は増加したが売上は減少/横ばい")
elif total_ad_change <= 0 and total_rev_change > 0:
    print("  → 広告費は減少したが売上は増加（広告以外の成長要因あり）")
else:
    print("  → 広告費・売上ともに減少")

# Monthly patterns
print(f"\n  月次パターン分析:")
for i in range(len(rev_chg)):
    ad_dir = "↑" if ad_chg[i] > 0 else "↓"
    rev_dir = "↑" if rev_chg[i] > 0 else "↓"
    print(f"    {MONTH_LABELS[i+1]}: 広告{ad_dir}({ad_chg[i]:+,.0f}) → 売上{rev_dir}({rev_chg[i]:+,.0f})")

# b. 広告費の増加なしでも売上は成長していたか？
print(f"\n■ b. 広告費の増加なしでも売上は成長していたか？")
# Check months where ad decreased but revenue changed
months_ad_down_rev_up = []
months_ad_down_rev_down = []
for i in range(len(rev_chg)):
    if ad_chg[i] < 0:
        if rev_chg[i] > 0:
            months_ad_down_rev_up.append(MONTH_LABELS[i+1])
        else:
            months_ad_down_rev_down.append(MONTH_LABELS[i+1])

print(f"  広告費減少→売上増加の月: {', '.join(months_ad_down_rev_up) if months_ad_down_rev_up else 'なし'}")
print(f"  広告費減少→売上減少の月: {', '.join(months_ad_down_rev_down) if months_ad_down_rev_down else 'なし'}")

# Calculate ad-independent growth indicator
# Revenue growth that can't be explained by ad spend changes
print(f"\n  広告費が下がっても売上が増加したケースが{'ある' if months_ad_down_rev_up else 'ない'}:")
if months_ad_down_rev_up:
    print(f"  → 広告費以外の成長ドライバー（例: ブランド認知蓄積、自然流入、商品力）が存在する可能性")
else:
    print(f"  → 売上は広告費に強く依存している可能性")

# c. 広告の限界効率は改善/悪化しているか？
print(f"\n■ c. 広告の限界効率（追加1円の広告で得られる追加売上）は改善/悪化しているか？")
print(f"  広告1円あたり売上（平均効率）の推移:")
for i, m in enumerate(MONTHS):
    rpa = rev_per_ad[i]
    if rpa is not None:
        print(f"    {MONTH_LABELS[i]}: {rpa:.2f}倍")

# Trend of average efficiency
if rev_per_ad[0] and rev_per_ad[-1]:
    if rev_per_ad[-1] < rev_per_ad[0]:
        print(f"  → 平均効率は悪化傾向 ({rev_per_ad[0]:.2f} → {rev_per_ad[-1]:.2f})")
        print(f"    広告費の増加ペースが売上増加ペースを上回っている")
    else:
        print(f"  → 平均効率は改善傾向 ({rev_per_ad[0]:.2f} → {rev_per_ad[-1]:.2f})")

print(f"\n  限界効率（Δ売上/Δ広告費）の推移:")
for i in range(1, len(MONTHS)):
    me = marginal_eff[i]
    if me is not None:
        status = "効率的" if me > 1 else ("非効率" if me > 0 else "逆効果")
        print(f"    {MONTH_LABELS[i]}: {me:.2f}倍 ({status})")

# d. どのセグメントで広告が効いていて、どこで効いていないか？
print(f"\n■ d. どのセグメントで広告が効いていて、どこで効いていないか？")

effective_segs = []
ineffective_segs = []

for seg in active_segments:
    revs_s = seg_data[seg]["rev"]
    ads_s = seg_data[seg]["ad"]

    # Calculate revenue per ad spend for first and last month
    rpa_first = revs_s[0] / ads_s[0] if ads_s[0] > 0 else None
    rpa_last = revs_s[-1] / ads_s[-1] if ads_s[-1] > 0 else None

    # MoM delta correlation
    rev_d = mom_change(revs_s)
    ad_d = mom_change(ads_s)
    corr = calc_corr(ad_d, rev_d)

    # Overall change
    rev_change_s = revs_s[-1] - revs_s[0]
    ad_change_s = ads_s[-1] - ads_s[0]

    # Flags for ad increase but revenue decrease
    ad_up_rev_down_count = sum(1 for i in range(len(rev_d)) if ad_d[i] > 0 and rev_d[i] <= 0)

    entry = {
        "name": seg,
        "rev_change": rev_change_s,
        "ad_change": ad_change_s,
        "corr": corr,
        "rpa_first": rpa_first,
        "rpa_last": rpa_last,
        "ad_up_rev_down": ad_up_rev_down_count,
        "total_ad": sum(ads_s),
        "total_rev": sum(revs_s)
    }

    # Classify: effective if correlation > 0.3 and ad increase led to revenue increase
    if corr is not None and corr > 0.3:
        effective_segs.append(entry)
    else:
        ineffective_segs.append(entry)

print("\n  【広告が効いているセグメント】（変動相関 > 0.3）")
header7 = "セグメント".ljust(30) + "相関r".rjust(10) + "8ヶ月広告費計".rjust(16) + "8ヶ月売上計".rjust(16) + "広告増→売上減回数".rjust(18)
print(f"  {header7}")
print(f"  {'-' * len(header7)}")
for e in sorted(effective_segs, key=lambda x: x["corr"] if x["corr"] else 0, reverse=True):
    print(f"  {e['name'].ljust(30)}{fmt_r(e['corr'], 10)}{fmt(e['total_ad'], 16)}{fmt(e['total_rev'], 16)}{str(e['ad_up_rev_down']).rjust(18)}")

print("\n  【広告効果が不明確/非効率なセグメント】（変動相関 <= 0.3）")
print(f"  {header7}")
print(f"  {'-' * len(header7)}")
for e in sorted(ineffective_segs, key=lambda x: x["total_ad"], reverse=True):
    print(f"  {e['name'].ljust(30)}{fmt_r(e['corr'], 10)}{fmt(e['total_ad'], 16)}{fmt(e['total_rev'], 16)}{str(e['ad_up_rev_down']).rjust(18)}")

# e. 「広告費を減らしたら売上はどうなるか」の示唆
print(f"\n■ e. 「広告費を減らしたら売上はどうなるか」の示唆")

# High-ad, low-correlation segments (waste candidates)
print("\n  【広告削減候補】高支出 & 低相関セグメント:")
waste_candidates = [e for e in ineffective_segs if e["total_ad"] > 1000000]
waste_candidates.sort(key=lambda x: x["total_ad"], reverse=True)

if waste_candidates:
    for e in waste_candidates:
        avg_ad = e["total_ad"] / 8
        avg_rev = e["total_rev"] / 8
        rpa = avg_rev / avg_ad if avg_ad > 0 else 0
        print(f"    {e['name']}: 月平均広告費 {avg_ad:,.0f}円, 月平均売上 {avg_rev:,.0f}円, 効率 {rpa:.1f}倍, 相関 r={e['corr']:.2f}" if e['corr'] else f"    {e['name']}: 月平均広告費 {avg_ad:,.0f}円, 相関計算不可")
else:
    print("    該当なし")

print("\n  【広告維持推奨】高支出 & 高相関セグメント:")
keep_candidates = [e for e in effective_segs if e["total_ad"] > 1000000]
keep_candidates.sort(key=lambda x: x["total_ad"], reverse=True)

if keep_candidates:
    for e in keep_candidates:
        avg_ad = e["total_ad"] / 8
        avg_rev = e["total_rev"] / 8
        rpa = avg_rev / avg_ad if avg_ad > 0 else 0
        print(f"    {e['name']}: 月平均広告費 {avg_ad:,.0f}円, 月平均売上 {avg_rev:,.0f}円, 効率 {rpa:.1f}倍, 相関 r={e['corr']:.2f}" if e['corr'] else f"    {e['name']}: 月平均広告費 {avg_ad:,.0f}円")

# Overall recommendation
print(f"\n■ 総合結論:")

# Determine dominant pattern
if corr_d_same and corr_d_same > 0.5:
    print(f"  1. 広告費と売上の同月変動に強い正の相関あり（r={corr_d_same:.2f}）")
    print(f"     → 広告費削減は短期的に売上減少リスクあり")
elif corr_d_same and corr_d_same > 0.3:
    print(f"  1. 広告費と売上の同月変動に中程度の正の相関あり（r={corr_d_same:.2f}）")
    print(f"     → 広告費と売上の連動は部分的")
else:
    corr_val = f"r={corr_d_same:.2f}" if corr_d_same else "N/A"
    print(f"  1. 広告費と売上の同月変動に明確な相関なし（{corr_val}）")
    print(f"     → 広告費以外の要因が売上を大きく左右している可能性")

if corr_d_lag1 and corr_d_lag1 > corr_d_same if corr_d_same else False:
    print(f"  2. 1ヶ月ラグ相関({corr_d_lag1:.2f}) > 同月相関({corr_d_same:.2f})")
    print(f"     → 広告効果は翌月に現れる傾向あり（投資回収に1ヶ月の遅延）")
else:
    lag_val = f"{corr_d_lag1:.2f}" if corr_d_lag1 else "N/A"
    same_val = f"{corr_d_same:.2f}" if corr_d_same else "N/A"
    print(f"  2. ラグ相関({lag_val}) vs 同月相関({same_val})")
    print(f"     → 広告効果は主に同月内で発現（即効性が高い）")

# Ad efficiency trend
if rev_per_ad[0] and rev_per_ad[-1]:
    if rev_per_ad[-1] < rev_per_ad[0]:
        print(f"  3. 広告効率は悪化傾向（{rev_per_ad[0]:.1f}倍→{rev_per_ad[-1]:.1f}倍）")
        print(f"     → 広告の限界効用逓減が発生している可能性（飽和状態に近づいている）")
    else:
        print(f"  3. 広告効率は改善傾向（{rev_per_ad[0]:.1f}倍→{rev_per_ad[-1]:.1f}倍）")

ineffective_ad_total = sum(e["total_ad"] for e in ineffective_segs)
total_all_ad = sum(data[m]["広告宣伝費"]["合計"] for m in MONTHS)
ineffective_pct = ineffective_ad_total / total_all_ad * 100 if total_all_ad > 0 else 0

print(f"  4. 広告効果が不明確なセグメントの広告費: 8ヶ月計 {ineffective_ad_total:,.0f}円（全体の{ineffective_pct:.1f}%）")
print(f"     → このセグメントの広告費最適化が改善余地")

print(f"\n  【提言】")
print(f"  ・効果不明確セグメントの広告費を段階的に削減し、売上変動をモニタリング")
print(f"  ・高相関セグメントへの広告費の再配分を検討")
print(f"  ・1ヶ月ラグを考慮した広告効果測定の仕組み構築を推奨")
print(f"  ・売上が広告費なしでも維持されるベースライン水準の把握が重要")

print("\n" + "=" * 100)
print("分析完了")
print("=" * 100)
